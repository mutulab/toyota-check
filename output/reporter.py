"""結果をExcel/CSVに出力"""

import csv
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
from pathlib import Path

RED   = PatternFill("solid", fgColor="FFCCCC")
YEL   = PatternFill("solid", fgColor="FFFF99")
GREEN = PatternFill("solid", fgColor="CCFFCC")
BOLD  = Font(bold=True)

OUTPUT_DIR = Path(__file__).parent.parent / "reports"


def _ensure_dir():
    OUTPUT_DIR.mkdir(exist_ok=True)
    return OUTPUT_DIR


def save_links_report(results: list[dict], filename: str = "") -> str:
    """リンクチェック結果をExcelに出力"""
    _ensure_dir()
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    path = OUTPUT_DIR / (filename or f"links_{ts}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "リンクチェック"

    headers = ["URL", "ソース", "HTTPステータス", "判定", "備考"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = BOLD

    for r in sorted(results, key=lambda x: (not x.get("broken", False), x["url"])):
        status = r.get("status", 0)
        ok = r.get("ok", False)
        broken = r.get("broken", False)
        row = [r["url"], r.get("source", ""), status,
               "❌ リンク切れ" if broken else ("✅ 正常" if ok else "⚠️ 要確認"),
               ""]
        ws.append(row)
        fill = RED if broken else (YEL if not ok else None)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    ws.column_dimensions["A"].width = 60
    wb.save(path)
    print(f"  → {path}")
    return str(path)


def save_cwv_report(results: list[dict], filename: str = "") -> str:
    """CWV結果をExcelに出力"""
    _ensure_dir()
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    path = OUTPUT_DIR / (filename or f"cwv_{ts}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Core Web Vitals"

    headers = [
        "URL", "スコア", "策略",
        "LCP(ms)", "LCP判定", "CLS", "CLS判定",
        "INP(ms)", "INP判定", "FCP(ms)", "TTFB(ms)", "NG項目"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = BOLD

    for r in results:
        if r.get("error"):
            ws.append([r["url"], "ERROR"] + [""] * 10)
            continue
        row = [
            r["url"], r.get("perf_score"), r.get("strategy"),
            r.get("LCP"), r.get("LCP_rate"),
            r.get("CLS"), r.get("CLS_rate"),
            r.get("INP"), r.get("INP_rate"),
            r.get("FCP"), r.get("TTFB"),
            ", ".join(r.get("ng_items", [])),
        ]
        ws.append(row)
        # 色付け
        ng = r.get("ng_items", [])
        fill = RED if ng else GREEN
        for cell in ws[ws.max_row]:
            cell.fill = fill

    ws.column_dimensions["A"].width = 60
    wb.save(path)
    print(f"  → {path}")
    return str(path)


def save_content_report(results: list[dict], filename: str = "") -> str:
    """コンテンツチェック結果をExcelに出力"""
    _ensure_dir()
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    path = OUTPUT_DIR / (filename or f"content_{ts}.xlsx")
    wb = openpyxl.Workbook()

    # シート1: 問題あり一覧
    ws1 = wb.active
    ws1.title = "要確認リスト"
    headers = ["URL", "title", "description", "title問題", "表記ゆれ", "禁止表現"]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = BOLD

    # シート2: 全件
    ws2 = wb.create_sheet("全件")
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = BOLD

    for r in results:
        if r.get("error"):
            continue
        yure = "; ".join(
            f"{y['matches']}→{y['correct']}" for y in r.get("hyoki_yure", [])
        )
        kinshi = "; ".join(
            f"{k['matches']}({k['reason']})" for k in r.get("kinshi_hyogen", [])
        )
        row = [
            r["url"], r.get("title", ""), r.get("description", ""),
            "; ".join(r.get("title_issues", [])), yure, kinshi,
        ]
        ws2.append(row)
        if r.get("ng"):
            ws1.append(row)
            for cell in ws1[ws1.max_row]:
                cell.fill = YEL

    for ws in [ws1, ws2]:
        ws.column_dimensions["A"].width = 55
        ws.column_dimensions["B"].width = 40
    wb.save(path)
    print(f"  → {path}")
    return str(path)


def update_excel_titles(excel_path: str, titles: dict[str, dict]) -> int:
    """tjpコンテンツ管理表のページ説明列をtitle/descで更新"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["運用サイトマップ"]

    from urllib.parse import urlparse
    import re

    def normalize(url):
        url = url.rstrip("/")
        if url.endswith("/index.html"):
            url = url[: -len("/index.html")]
        return url

    # 正規化済みtitlesマップ
    norm_titles = {normalize(u): m for u, m in titles.items() if m}

    updated = 0
    for row_num in range(6, ws.max_row + 1):
        url = ws.cell(row_num, 8).value
        if not url or not isinstance(url, str):
            continue
        meta = norm_titles.get(normalize(url.strip()))
        if meta and meta.get("short_title"):
            desc = meta["short_title"]
            if meta.get("description"):
                desc += f" | {meta['description'][:60]}"
            ws.cell(row_num, 9).value = desc
            updated += 1

    wb.save(excel_path)
    return updated
