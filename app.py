"""toyota-check Streamlit Web アプリ"""

import io
import sys
import time
import pandas as pd
import streamlit as st
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

# ─── ページ設定 ───────────────────────────────────────
st.set_page_config(
    page_title="toyota-check",
    page_icon="🚗",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── パスワード認証 ───────────────────────────────────
def check_password() -> bool:
    if st.session_state.get("authenticated"):
        return True

    st.title("🚗 toyota-check")
    st.caption("toyota.jp 専用サイト検証ツール")
    st.divider()

    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        pw = st.text_input("パスワード", type="password", key="pw_input",
                           placeholder="チーム共有パスワードを入力")
        if st.button("ログイン", use_container_width=True, type="primary"):
            correct = st.secrets.get("PASSWORD", "toyota-tqp-2026")
            if pw == correct:
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("パスワードが違います")
    return False

if not check_password():
    st.stop()

# ─── インポート（認証後） ────────────────────────────
from fetcher import fetch_html, extract_meta
from config import CONCURRENT_WORKERS, REQUEST_DELAY

# ─── ヘルパー ─────────────────────────────────────────
# コンテンツ管理票のシート別読み込み設定（col は 0始まりの列番号）
# filter_col/filter_substr: 指定列にその文字列を含む行のみ採用
SHEET_CONFIGS = {
    "運用サイトマップ":                    {"min_row": 6, "col": 7},
    "コンテンツ_車種":                     {"min_row": 5, "col": 2},
    "コンテンツ_車種横断":                 {"min_row": 3, "col": 2},
    "TQP非検知info・informationフォルダ":  {"min_row": 2, "col": 1},
    # 対応列（B列）が「移行」を含む行のみ（使用確認・削除・移設等は除外）
    "CMS移行":                             {"min_row": 2, "col": 0,
                                            "filter_col": 1, "filter_substr": "移行"},
}

def load_urls_from_excel(uploaded_file, sheets: list[str] | None = None,
                         show_counts: bool = True) -> list[str]:
    """コンテンツ管理票から対象シートのURLを収集して返す。

    相対パス（/info/... 等）は https://toyota.jp を補完。シート横断で重複排除。
    """
    import openpyxl
    wb = openpyxl.load_workbook(uploaded_file, read_only=True)
    targets = [s for s in (sheets or list(SHEET_CONFIGS)) if s in wb.sheetnames]
    if not targets:
        st.error("対象シートが見つかりません（運用サイトマップ / コンテンツ_車種 等）")
        return []

    urls: list[str] = []
    counts: dict[str, int] = {}
    for sn in targets:
        cfg = SHEET_CONFIGS[sn]
        before = len(urls)
        for row in wb[sn].iter_rows(min_row=cfg["min_row"], values_only=True):
            if cfg["col"] >= len(row):
                continue
            u = row[cfg["col"]]
            if not u or not isinstance(u, str):
                continue
            fc = cfg.get("filter_col")
            if fc is not None:
                fv = row[fc] if fc < len(row) else None
                if not fv or cfg["filter_substr"] not in str(fv):
                    continue
            u = u.strip()
            if u.startswith("/"):
                u = "https://toyota.jp" + u
            if u.startswith("http"):
                urls.append(u)
        counts[sn] = len(urls) - before

    deduped = list(dict.fromkeys(urls))
    if show_counts and len(targets) > 1:
        st.caption("　/　".join(f"{sn}: {c}件" for sn, c in counts.items())
                   + f"　→ 重複排除後 {len(deduped)} 件")
    return deduped

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    return buf.getvalue()

# ─── UI ──────────────────────────────────────────────
st.title("🚗 toyota-check")
st.caption("toyota.jp 専用サイト検証ツール")

# サイドバー
with st.sidebar:
    st.header("設定")

    check_type = st.radio(
        "チェック種別",
        ["📄 タイトル取得", "🔗 リンクチェック", "🌐 外部リンクチェック",
         "📌 リンク元調査", "⚡ Core Web Vitals", "📝 表記ゆれ・禁止表現",
         "🔍 アプリ・機能検出", "🗂️ サイトマップ管理",
         "🕷️ バックグラウンドクロール", "📖 マニュアル"],
        index=0,
    )

    # ── バックグラウンドクロール専用設定 ──────────────────────────────
    bg_cfg: dict = {}
    if check_type == "🕷️ バックグラウンドクロール":
        st.divider()
        st.subheader("ジョブ設定")

        _bg_url_src = st.radio(
            "URL取得方法",
            ["🕷️ クロール（自動収集）", "📊 コンテンツ管理票（Excel）"],
            key="bg_url_src",
        )
        if _bg_url_src.startswith("🕷"):
            bg_cfg["url_source_type"] = "crawl"
            bg_cfg["start_url"] = st.text_input("開始URL", value="https://toyota.jp/")
            bg_cfg["max_pages"] = st.slider("最大収集ページ数", 10, 1000, 100)
            bg_cfg["depth"]     = int(st.number_input("クロール深さ", 1, 4, 2))
            bg_cfg["path_filter"] = st.text_input(
                "パス制限（任意）",
                value="",
                placeholder="/prius/",
                help="指定したパスで始まるURLのみ収集・巡回します。空欄=ドメイン全体",
                key="bg_path_filter",
            ).strip()
        else:
            bg_cfg["url_source_type"] = "excel"
            bg_cfg["start_url"] = ""
            _bg_exc = st.file_uploader(
                "コンテンツ管理票.xlsx", type=["xlsx"], key="bg_exc_ul"
            )
            _bg_sheets = st.multiselect(
                "読み込むシート",
                list(SHEET_CONFIGS),
                default=list(SHEET_CONFIGS),
                key="bg_exc_sheets",
            )
            if _bg_exc and _bg_sheets:
                _fk = f"{_bg_exc.name}_{_bg_exc.size}_{'|'.join(_bg_sheets)}"
                if st.session_state.get("_bg_exc_key") != _fk:
                    _exc_urls = load_urls_from_excel(_bg_exc, _bg_sheets)
                    st.session_state["bg_excel_urls"] = _exc_urls
                    st.session_state["_bg_exc_key"] = _fk
            bg_cfg["urls"]      = st.session_state.get("bg_excel_urls", [])
            bg_cfg["max_pages"] = len(bg_cfg["urls"])
            bg_cfg["depth"]     = 0
            if bg_cfg["urls"]:
                st.caption(f"✅ {len(bg_cfg['urls'])} 件のURLを読み込みました")
            else:
                st.info("コンテンツ管理票をアップロードしてください")

        bg_cfg["toyota_only"]  = st.checkbox("toyota.jpリンクのみ", value=True)
        _bg_checks = st.multiselect(
            "実行するチェック",
            ["リンクチェック", "外部リンクチェック", "リンク元調査",
             "表記ゆれ・禁止表現", "Core Web Vitals", "アプリ・機能検出"],
            default=["リンクチェック"],
        )
        _map = {"リンクチェック": "link", "外部リンクチェック": "extlink",
                "リンク元調査": "backlink", "表記ゆれ・禁止表現": "content",
                "Core Web Vitals": "cwv", "アプリ・機能検出": "app"}
        bg_cfg["check_types"] = [_map[c] for c in _bg_checks]
        if "link" in bg_cfg["check_types"]:
            st.caption("リンクチェック: リソース種別")
            c1, c2 = st.columns(2)
            _sel: set = set()
            if c1.checkbox("リンク(<a>)",   value=True,  key="bg_rl"): _sel.add("リンク")
            if c1.checkbox("CSS",           value=True,  key="bg_rc"): _sel.add("CSS/スタイル")
            if c1.checkbox("JavaScript",    value=True,  key="bg_rj"): _sel.add("JavaScript")
            if c2.checkbox("画像",          value=True,  key="bg_ri"): _sel.update({"画像", "画像(インラインCSS)"})
            if c2.checkbox("メディア",      value=False, key="bg_rm"): _sel.add("メディア")
            if c2.checkbox("iframe",        value=False, key="bg_rf"): _sel.add("iframe")
            if st.checkbox("CSS内リソース", value=False, key="bg_rcr"): _sel.update({"画像(CSS)", "フォント(CSS)", "CSSリソース"})
            bg_cfg["selected_res"] = list(_sel)
        if "content" in bg_cfg["check_types"]:
            bg_cfg["custom_dict"] = st.text_area(
                "カスタム辞書（誤表記|推奨表記）", height=80,
                placeholder="ウエブサイト|WEBサイト", key="bg_dict",
            )
        if "backlink" in bg_cfg["check_types"]:
            bg_cfg["backlink_query"] = st.text_input(
                "調査対象URL（部分一致）",
                placeholder="/lecture/ または https://factory.kinto-jp.com/",
                help="このURLパターンへのリンクを含むページを抽出します",
                key="bg_bl_query",
            ).strip()
        if "cwv" in bg_cfg["check_types"]:
            bg_cfg["strategy"] = st.radio("CWVデバイス", ["mobile", "desktop"],
                                          horizontal=True, key="bg_strat")
            bg_cfg["psi_key"]  = st.text_input("PSI API Key", type="password",
                                               value=st.secrets.get("PSI_API_KEY", ""),
                                               key="bg_psi")
        # URL source 変数をデフォルト値で初期化（後続コードで参照されるため）
        url_source = None
        uploaded_file = None
        manual_urls: list = []
        crawl_start = ""
        crawl_max = 100
        crawl_depth = 2
        limit = 0
        toyota_only = False
        selected_res_types: set = set()
        custom_dict_raw = ""
        strategy = "mobile"
        psi_key = st.secrets.get("PSI_API_KEY", "")

    elif check_type == "🗂️ サイトマップ管理":
        # ── サイトマップ管理: URLソースUIは不要（本体側で完結） ──────────
        st.divider()
        st.caption("メイン画面でコンテンツ管理表をアップロードしてください。")
        url_source = None
        uploaded_file = None
        manual_urls = []
        crawl_start = ""
        crawl_max = 100
        crawl_depth = 2
        limit = 0
        toyota_only = False
        selected_res_types: set = set()
        custom_dict_raw = ""
        strategy = "mobile"
        psi_key = st.secrets.get("PSI_API_KEY", "")

    else:
        # ── 通常モード: URL ソース ──────────────────────────────────────
        st.divider()
        st.subheader("URL ソース")
        url_source = st.radio(
            "",
            ["Excelファイルをアップロード", "URLを直接入力", "🕷️ クロール（自動収集）"],
            label_visibility="collapsed",
        )

        uploaded_file = None
        manual_urls = []
        crawl_start = ""
        crawl_max = 100
        crawl_depth = 2
        crawl_path_filter = ""

        if url_source == "Excelファイルをアップロード":
            uploaded_file = st.file_uploader(
                "tjpコンテンツ管理表.xlsx",
                type=["xlsx"],
                help="運用サイトマップ / コンテンツ_車種 / 車種横断 / TQP非検知 / CMS移行 の各シートからURLを読み込みます",
            )
            excel_sheets = st.multiselect(
                "読み込むシート",
                list(SHEET_CONFIGS),
                default=list(SHEET_CONFIGS),
                key="norm_exc_sheets",
            )
            limit = st.slider("上限URL数（0=全件）", 0, 200, 0,
                              help="テスト時は10〜20程度に設定推奨")
        elif url_source == "URLを直接入力":
            raw = st.text_area("URLを1行ずつ入力", height=200,
                               placeholder="https://toyota.jp/alphard/\nhttps://toyota.jp/prius/")
            manual_urls = [u.strip() for u in raw.splitlines() if u.strip().startswith("http")]
            limit = 0
        else:  # クロール（自動収集）
            crawl_start = st.text_input("開始URL", value="https://toyota.jp/")
            crawl_max = st.slider("最大収集ページ数", 10, 1000, 100,
                                  help="まず50前後で試してください")
            crawl_depth = st.number_input("クロール深さ（階層数）", min_value=1, max_value=4, value=2,
                                          help="1=直リンクのみ / 2=その先も辿る")
            crawl_path_filter = st.text_input(
                "パス制限（任意）",
                value="",
                placeholder="/prius/",
                help="指定したパスで始まるURLのみ収集・巡回します。空欄=ドメイン全体",
            ).strip()
            limit = 0

        # ── リンクチェック: リソース種別選択 ───────────────────────────
        if check_type == "🔗 リンクチェック":
            toyota_only = st.checkbox(
                "toyota.jpリンクのみ",
                value=True,
                help="toyota.jpドメイン以外を除外（高速化）",
            )
            st.divider()
            st.caption("チェック対象リソース種別")
            c1, c2 = st.columns(2)
            _sel2: set = set()
            if c1.checkbox("リンク(<a>)",  value=True,  key="rl"): _sel2.add("リンク")
            if c1.checkbox("CSS",          value=True,  key="rc"): _sel2.add("CSS/スタイル")
            if c1.checkbox("JavaScript",   value=True,  key="rj"): _sel2.add("JavaScript")
            if c2.checkbox("画像",         value=True,  key="ri"): _sel2.update({"画像", "画像(インラインCSS)"})
            if c2.checkbox("メディア",     value=False, key="rm"): _sel2.add("メディア")
            if c2.checkbox("iframe",       value=False, key="rf"): _sel2.add("iframe")
            if st.checkbox("CSS内リソース（背景画像・フォント）", value=False, key="rcr"):
                _sel2.update({"画像(CSS)", "フォント(CSS)", "CSSリソース"})
            selected_res_types = _sel2
        else:
            toyota_only = False
            selected_res_types = set()

        if check_type == "📝 表記ゆれ・禁止表現":
            st.divider()
            st.subheader("表記ゆれ辞書")
            from dict_loader import load_flat, source_label as _src_label
            from pathlib import Path as _Path
            _TMP_D     = _Path("/tmp/toyota-check-hyoki.xlsx")
            _BUNDLED_D = _Path(__file__).parent / "hyoki_dict.xlsx"
            _cur_flat  = load_flat()
            _using_tmp = _TMP_D.exists()
            st.caption(f"{_src_label()} · {len(_cur_flat)} 件")

            # 現在の辞書をダウンロード
            _dl_path = _TMP_D if _using_tmp else (_BUNDLED_D if _BUNDLED_D.exists() else None)
            if _dl_path:
                st.download_button(
                    "📥 現在の辞書をダウンロード",
                    _dl_path.read_bytes(),
                    file_name="hyoki_dict.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_cur_dict",
                )
            if _using_tmp:
                st.warning(
                    "セッション辞書を使用中です。\n\n"
                    "アプリ再起動で消えるため、永続化するには↑でダウンロードして"
                    " `hyoki_dict.xlsx` としてリポジトリに commit してください。",
                    icon="⚠️",
                )

            with st.expander("辞書を更新（Excelアップロード）"):
                _dict_xl = st.file_uploader(
                    "toyota_lexus用語リスト.xlsx", type=["xlsx"], key="dict_xl_up"
                )
                if _dict_xl:
                    _dk = f"{_dict_xl.name}_{_dict_xl.size}"
                    if st.session_state.get("_dict_xl_key") != _dk:
                        st.session_state["_dict_new_bytes"] = _dict_xl.read()
                        st.session_state["_dict_xl_key"] = _dk
            custom_dict_raw = st.text_area(
                "追加エントリ（1行1件: NG表記|推奨表記）",
                height=100,
                placeholder="ウエブサイト|WEBサイト\nお問合せ|お問い合わせ",
                help="辞書にない追加チェック項目。正規表現も使用可。",
            )
        else:
            custom_dict_raw = ""

        if check_type == "📌 リンク元調査":
            st.divider()
            st.subheader("調査設定")
            backlink_query = st.text_input(
                "調査対象URL（部分一致）",
                placeholder="/lecture/",
                help="このURLパターンへのリンクを含むページを抽出します。\n"
                     "例: /lecture/  →  href に /lecture/ を含むリンクを検出\n"
                     "例: https://factory.kinto-jp.com/  →  外部リンクも検出",
            ).strip()
        else:
            backlink_query = ""

        if check_type == "⚡ Core Web Vitals":
            strategy = st.radio("計測デバイス", ["mobile", "desktop"], horizontal=True)
            psi_key = st.text_input("PSI API Key（任意）",
                                    value=st.secrets.get("PSI_API_KEY", ""),
                                    type="password",
                                    help="未入力でも動作します（25,000回/日の上限あり）")
        else:
            strategy = "mobile"
            psi_key = st.secrets.get("PSI_API_KEY", "")

    st.divider()
    if check_type == "🗂️ サイトマップ管理":
        run_btn = False
    else:
        run_btn = st.button("▶ チェック実行", type="primary", use_container_width=True)

# ─── サイトマップ管理 UI（常時表示 / 通常フローをスキップ） ──────────
if check_type == "🗂️ サイトマップ管理":
    try:
        import sitemap_manager as _sm
        _sm.render()
    except Exception as _e:
        st.error(f"サイトマップ管理の読み込みに失敗しました: {_e}")
    st.stop()

# ─── バックグラウンドクロール UI（常時表示 / 通常フローをスキップ） ──
if check_type == "🕷️ バックグラウンドクロール":
    try:
        import crawler as _cw
    except Exception as _e:
        st.error(f"crawler モジュールの読み込みに失敗しました: {_e}")
        st.stop()

    tab_new, tab_status = st.tabs(["▶ 新規ジョブ開始", "🔍 ジョブ確認"])

    with tab_new:
        st.subheader("ジョブ設定")
        st.json({k: v for k, v in bg_cfg.items() if k != "psi_key"})
        if run_btn:  # サイドバーの「▶ チェック実行」= ジョブ開始
            if bg_cfg.get("url_source_type") == "excel" and not bg_cfg.get("urls"):
                st.error("コンテンツ管理票のExcelをアップロードしてください。")
            else:
                try:
                    jid = _cw.start_job(bg_cfg)
                    st.session_state["bg_job_id"] = jid
                    st.success(f"ジョブを開始しました。**ジョブID: `{jid}`**")
                    st.info("ブラウザを閉じても処理は継続します。「🔍 ジョブ確認」タブで結果を確認してください。")
                except Exception as _e:
                    st.error(f"ジョブ開始に失敗しました: {_e}")
        else:
            st.info("設定を確認して、サイドバーの「▶ チェック実行」を押すとジョブが開始されます。")

    with tab_status:
        # ── ジョブ選択（履歴 or 直接入力） ───────────────────────────
        recent = _cw.list_jobs()
        job_id_input = ""

        if recent:
            _NONE = "— IDを直接入力 —"
            def _jlabel(j):
                src = j["cfg"].get("start_url") or f"Excel ({j['cfg'].get('max_pages', '?')}件)"
                return f"{j['id']}  [{j['status']}]  {src}  ({j['started_at'][:16]})"
            _opts: dict = {_jlabel(j): j["id"] for j in recent}
            _all_keys = [_NONE] + list(_opts.keys())
            _sess_id = st.session_state.get("bg_job_id", "")
            _def_key = next((k for k, v in _opts.items() if v == _sess_id), _NONE)
            _sel = st.selectbox(
                "📋 ジョブ履歴から選択",
                _all_keys,
                index=_all_keys.index(_def_key),
            )
            if _sel != _NONE:
                job_id_input = _opts[_sel]

        if not job_id_input:
            job_id_input = st.text_input(
                "ジョブIDを直接入力",
                value=st.session_state.get("bg_job_id", ""),
                placeholder="例: A1B2C3D4",
            ).strip().upper()

        # ── ジョブ詳細 ────────────────────────────────────────────────
        if job_id_input:
            job = _cw.get_job(job_id_input)
            if not job:
                st.error("ジョブが見つかりません。IDを確認してください。")
            else:
                st.progress(job["progress"] / 100, text=job["phase"])
                col1, col2, col3 = st.columns(3)
                col1.metric("ステータス", job["status"])
                col2.metric("収集ページ数", job.get("url_count", 0))
                col3.metric("進捗", f"{job['progress']}%")
                st.caption(f"開始: {job['started_at']}　完了: {job.get('finished_at') or '—'}")

                if job["status"] == "error":
                    st.error(f"エラー: {job.get('error', '')}")

                if job["status"] == "running":
                    from datetime import datetime as _dt
                    _last = job.get("last_updated_at") or job.get("started_at", "")
                    try:
                        _elapsed = (_dt.now() - _dt.fromisoformat(_last)).total_seconds()
                        if _elapsed > 600:
                            st.warning(
                                f"⚠️ 最終更新から {int(_elapsed / 60)} 分経過しています。"
                                "タスクがスタックしている可能性があります。"
                            )
                    except Exception:
                        pass
                    st.info("実行中です。完了したら「🔄 更新」を押してください。")
                    if st.button("🔄 更新"):
                        st.rerun()

                if job["status"] == "done":
                    res = job.get("results", {})
                    if "link" in res:
                        df_l = pd.DataFrame(res["link"]) if res["link"] else pd.DataFrame()
                        _broken_cnt = int((df_l["判定"] == "❌ 切れ").sum()) if not df_l.empty and "判定" in df_l.columns else len(df_l)
                        _unv_cnt    = int((df_l["判定"] == "⚠️ 確認不可").sum()) if not df_l.empty and "判定" in df_l.columns else 0
                        st.subheader(f"🔗 リンク切れ {_broken_cnt} 件 / 確認不可 {_unv_cnt} 件")
                        if not df_l.empty:
                            _df_broken_bg = df_l[df_l["判定"] == "❌ 切れ"] if "判定" in df_l.columns else df_l
                            _df_unv_bg    = df_l[df_l["判定"] == "⚠️ 確認不可"] if "判定" in df_l.columns else pd.DataFrame()
                            if not _df_broken_bg.empty:
                                st.write("**❌ リンク切れ（404/410）**")
                                st.dataframe(_df_broken_bg, use_container_width=True, height=250)
                            if not _df_unv_bg.empty:
                                with st.expander(f"⚠️ 確認不可 {len(_df_unv_bg)} 件"):
                                    st.caption("接続失敗（status=0）。実際にはリンク切れでない場合があります。")
                                    st.dataframe(_df_unv_bg, use_container_width=True, height=200)
                            c1, c2 = st.columns(2)
                            c1.download_button("📥 Excel", to_excel_bytes(df_l),
                                               f"links_{job_id_input}.xlsx", key="bg_dl_lx")
                            c2.download_button("📥 CSV",
                                               df_l.to_csv(index=False).encode("utf-8-sig"),
                                               f"links_{job_id_input}.csv", mime="text/csv",
                                               key="bg_dl_lc")
                    if "extlink" in res:
                        df_ex = pd.DataFrame(res["extlink"]) if res["extlink"] else pd.DataFrame()
                        _ex_broken = int((df_ex["判定"] == "❌ 切れ").sum()) if not df_ex.empty and "判定" in df_ex.columns else 0
                        _ex_unv    = int((df_ex["判定"] == "⚠️ 確認不可").sum()) if not df_ex.empty and "判定" in df_ex.columns else 0
                        st.subheader(f"🌐 外部リンク {len(df_ex)} 件 / 切れ {_ex_broken} 件 / 確認不可 {_ex_unv} 件")
                        if not df_ex.empty:
                            _df_ex_broken = df_ex[df_ex["判定"] == "❌ 切れ"] if "判定" in df_ex.columns else df_ex
                            _df_ex_unv    = df_ex[df_ex["判定"] == "⚠️ 確認不可"] if "判定" in df_ex.columns else pd.DataFrame()
                            if not _df_ex_broken.empty:
                                st.write("**❌ 外部リンク切れ（404/410）**")
                                st.dataframe(_df_ex_broken, use_container_width=True, height=200)
                            if not _df_ex_unv.empty:
                                with st.expander(f"⚠️ 確認不可 {len(_df_ex_unv)} 件"):
                                    st.dataframe(_df_ex_unv, use_container_width=True, height=200)
                            with st.expander(f"全外部リンク一覧 ({len(df_ex)} 件)"):
                                st.dataframe(df_ex, use_container_width=True, height=300)
                            st.download_button("📥 外部リンク Excel", to_excel_bytes(df_ex),
                                               f"extlinks_{job_id_input}.xlsx", key="bg_dl_ex")

                    if "backlink" in res:
                        df_bl = pd.DataFrame(res["backlink"]) if res["backlink"] else pd.DataFrame()
                        _bl_q = job["cfg"].get("backlink_query", "")
                        _bl_pages = len(df_bl["発見ページURL"].unique()) if not df_bl.empty and "発見ページURL" in df_bl.columns else 0
                        st.subheader(f"📌 リンク元調査「{_bl_q}」— {_bl_pages} ページで検出")
                        if not df_bl.empty:
                            st.dataframe(df_bl, use_container_width=True, height=300)
                            c1, c2 = st.columns(2)
                            c1.download_button("📥 Excel", to_excel_bytes(df_bl),
                                               f"backlinks_{job_id_input}.xlsx", key="bg_dl_blx")
                            c2.download_button("📥 CSV",
                                               df_bl.to_csv(index=False).encode("utf-8-sig"),
                                               f"backlinks_{job_id_input}.csv", mime="text/csv",
                                               key="bg_dl_blc")
                        else:
                            st.info(f"「{_bl_q}」へのリンクは見つかりませんでした。")

                    if "content" in res:
                        df_c = pd.DataFrame(res["content"]) if res["content"] else pd.DataFrame()
                        st.subheader(f"📝 表記ゆれ・禁止表現 ({len(df_c)} 件)")
                        if not df_c.empty:
                            st.dataframe(df_c, use_container_width=True, height=300)
                            c1, c2 = st.columns(2)
                            c1.download_button("📥 指示書 Excel", to_excel_bytes(df_c),
                                               f"content_{job_id_input}.xlsx", key="bg_dl_cx")
                            c2.download_button("📥 指示書 CSV",
                                               df_c.to_csv(index=False).encode("utf-8-sig"),
                                               f"content_{job_id_input}.csv", mime="text/csv",
                                               key="bg_dl_cc")
                    if "cwv" in res:
                        df_w = pd.DataFrame(res["cwv"]) if res["cwv"] else pd.DataFrame()
                        st.subheader(f"⚡ Core Web Vitals ({len(df_w)} 件)")
                        if not df_w.empty:
                            st.dataframe(df_w, use_container_width=True, height=300)
                            st.download_button("📥 CWV Excel", to_excel_bytes(df_w),
                                               f"cwv_{job_id_input}.xlsx", key="bg_dl_wx")
                    if "app" in res:
                        df_a = pd.DataFrame(res["app"]) if res["app"] else pd.DataFrame()
                        app_cnt = len(df_a[df_a["アプリ性スコア"] > 0]) if not df_a.empty else 0
                        st.subheader(f"🔍 アプリ・機能検出 (機能あり: {app_cnt} / {len(df_a)} ページ)")
                        if not df_a.empty:
                            st.dataframe(df_a, use_container_width=True, height=300)
                            st.download_button("📥 アプリ検出 Excel", to_excel_bytes(df_a),
                                               f"app_{job_id_input}.xlsx", key="bg_dl_ax")
    st.stop()

# ─── 表記ゆれ辞書差分確認（常時表示） ────────────────────────────────
if check_type == "📝 表記ゆれ・禁止表現" and "_dict_new_bytes" in st.session_state:
    import io as _io
    from dict_loader import load_flat, read_excel_flat, save_override
    try:
        _new_flat = read_excel_flat(_io.BytesIO(st.session_state["_dict_new_bytes"]))
        _cur_flat = load_flat()
        _added   = {k: v for k, v in _new_flat.items() if k not in _cur_flat}
        _removed = {k: v for k, v in _cur_flat.items() if k not in _new_flat}
        _changed = {k: (_cur_flat[k], v) for k, v in _new_flat.items()
                    if k in _cur_flat and _cur_flat[k] != v}
        total_delta = len(_added) + len(_removed) + len(_changed)
        if total_delta:
            st.subheader(f"📋 辞書差分プレビュー（+{len(_added)} / -{len(_removed)} / ~{len(_changed)}）")
            if _added:
                with st.expander(f"✅ 追加: {len(_added)} 件"):
                    st.dataframe(
                        pd.DataFrame([{"NG表記": k, "推奨表記": v} for k, v in _added.items()]),
                        use_container_width=True,
                    )
            if _removed:
                with st.expander(f"❌ 削除: {len(_removed)} 件"):
                    st.dataframe(
                        pd.DataFrame([{"NG表記": k, "推奨表記": v} for k, v in _removed.items()]),
                        use_container_width=True,
                    )
            if _changed:
                with st.expander(f"✏️ 変更: {len(_changed)} 件"):
                    st.dataframe(
                        pd.DataFrame([{"NG表記": k, "旧推奨": old, "新推奨": new}
                                      for k, (old, new) in _changed.items()]),
                        use_container_width=True,
                    )
            _ca, _cb = st.columns(2)
            if _ca.button("✅ 差分を適用", type="primary", key="apply_dict"):
                save_override(st.session_state["_dict_new_bytes"])
                st.session_state["_dict_applied_bytes"] = st.session_state["_dict_new_bytes"]
                for _k in ("_dict_new_bytes", "_dict_xl_key"):
                    st.session_state.pop(_k, None)
                st.success("辞書をセッションに適用しました。")
                st.rerun()
            if _cb.button("❌ キャンセル", key="cancel_dict"):
                for _k in ("_dict_new_bytes", "_dict_xl_key"):
                    st.session_state.pop(_k, None)
                st.rerun()
        else:
            st.info("アップロードした辞書と現在の辞書に差分はありません。")
            for _k in ("_dict_new_bytes", "_dict_xl_key"):
                st.session_state.pop(_k, None)
    except Exception as _dict_err:
        st.error(f"辞書の読み込みに失敗しました: {_dict_err}")
        for _k in ("_dict_new_bytes", "_dict_xl_key"):
            st.session_state.pop(_k, None)

# 差分適用直後: ダウンロード＆永続化案内を表示
if (check_type == "📝 表記ゆれ・禁止表現"
        and "_dict_applied_bytes" in st.session_state):
    st.success("✅ 辞書をセッションに適用しました。")
    st.info(
        "**永続化するには** 👇 からダウンロードして、"
        " `hyoki_dict.xlsx` としてリポジトリに commit してください。"
        "次回デプロイ以降、アップロード不要でこの辞書が使われます。"
    )
    st.download_button(
        "📥 適用済み辞書をダウンロード（hyoki_dict.xlsx）",
        st.session_state["_dict_applied_bytes"],
        file_name="hyoki_dict.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="dl_applied_dict",
        type="primary",
    )
    if st.button("閉じる", key="close_applied_banner"):
        st.session_state.pop("_dict_applied_bytes", None)
        st.rerun()

# ─── マニュアル（ボタン不要・即表示） ─────────────────────────────
if check_type == "📖 マニュアル":
    st.header("📖 toyota-check マニュアル")
    st.caption("toyota.jp 専用サイト検証ツール — 機能ガイド")

    st.info(
        "**基本的な使い方:** サイドバーでチェック種別と URL ソースを選び、"
        "「▶ チェック実行」を押すと結果が表示されます。"
    )

    # ── URL ソース ──────────────────────────────────────────────
    with st.expander("### 🗂️ URL ソースの選び方", expanded=True):
        st.markdown("""
| ソース | 向いているケース |
|--------|----------------|
| **Excelファイルをアップロード** | コンテンツ管理票（tjpコンテンツ管理表.xlsx）の各シートからURLを読み込む。「読み込むシート」で対象を選択、上限URL数で件数を絞れる |
| **URLを直接入力** | 特定のページだけ素早く確認したいとき。1行1URL で貼り付け |
| **🕷️ クロール** | 開始URLから自動でリンクを辿る。パス制限を使えば特定フォルダ（例: `/prius/`）に限定可能 |

**クロールのパス制限:**
- 開始URL: `https://toyota.jp/` + パス制限: `/prius/` → `/prius/` 以下のページのみ収集
- 空欄にするとドメイン全体を巡回
        """)

    # ── 各チェック種別 ─────────────────────────────────────────
    with st.expander("### 📄 タイトル取得"):
        st.markdown("""
各ページの **title タグ・description・H1** を一括取得します。

- `✅` — タイトルあり・正常取得
- `⚠️ titleなし` — title タグが空
- `❌` — ページ取得失敗（404 等）

**用途:** タイトル抜けの確認、SEO棚卸し
        """)

    with st.expander("### 🔗 リンクチェック"):
        st.markdown("""
各ページの **リンク・CSS・JS・画像・メディア** を収集し、ステータスを確認します。

**判定基準:**

| 判定 | 意味 | 対応 |
|------|------|------|
| ❌ 切れ | 404 / 410 | 確実なリンク切れ。修正要 |
| ⚠️ 確認不可 | status=0 | チェックサーバーからのアクセス制限・タイムアウト。ブラウザで個別確認 |
| ✅ 正常 | 200 系 | 問題なし |

**チェック対象リソース種別:**
- `リンク (<a>)` — ナビゲーションリンク
- `CSS` — スタイルシート
- `JavaScript` — JSファイル
- `画像` — img src・インラインCSS の url()
- `メディア` — video / audio / source
- `iframe` — 埋め込みフレーム
- `CSS内リソース` — CSSファイル内の背景画像・フォント（処理が重くなるため必要時のみ ON）

**toyota.jpリンクのみ オプション:** toyota.jp ドメイン以外のリソースを除外して高速化します。
        """)

    with st.expander("### 🌐 外部リンクチェック"):
        st.markdown("""
toyota.jp **以外のドメイン** への `<a href>` リンクを収集し、ステータスを確認します。

**判定基準:**

| 判定 | 意味 |
|------|------|
| ❌ 切れ | 404 / 410 — 外部サイトが削除・移転 |
| 🔒 アクセス制限 | 403 — チェックサーバーからのアクセス拒否（ブラウザでは開ける場合が多い） |
| ⚠️ 確認不可 | status=0 — タイムアウト・接続拒否 |
| ✅ 正常 | 200 系 |

**出力:** ドメイン別集計 + フィルタ付き全件一覧
        """)

    with st.expander("### 📌 リンク元調査（逆引き）"):
        st.markdown("""
**「このページへのリンクがどこにあるか」** を全ページスキャンして探します。

**用途例:**
- あるページが削除・移転した → 導線（リンク元）を洗い出して修正
- 外部サービスへのリンクがどこに貼られているか確認

**調査対象URL（部分一致）の書き方:**

| 入力例 | 検出対象 |
|--------|---------|
| `/lecture/` | `/lecture/` を含む href が貼られているページ |
| `https://factory.kinto-jp.com/` | 外部ドメインへのリンクを含むページ |
| `/prius` | `/prius/` `/priusPHEV/` など prius を含む全パス |

**出力:**
- ページ単位サマリ（ページURL / タイトル / リンク数）
- 詳細（どのリンクがマッチしたか）
- Excel / CSV ダウンロード
        """)

    with st.expander("### ⚡ Core Web Vitals"):
        st.markdown("""
Google PageSpeed Insights API を使い **LCP / CLS / INP** を計測します。

**KPI基準（TQP運用設計書準拠）:**

| 指標 | GOOD | NI | POOR |
|------|------|----|------|
| LCP | ≤ 2,500ms | ≤ 4,000ms | > 4,000ms |
| CLS | ≤ 0.1 | ≤ 0.25 | > 0.25 |
| INP | ≤ 200ms | ≤ 500ms | > 500ms |

- **デバイス:** mobile（推奨）/ desktop を切り替え可能
- **PSI API Key:** 未入力でも動作します（無料枠: 25,000回/日）。大量計測時はキーを取得推奨
- 計測には1ページあたり数秒かかります。大量URLの場合はバックグラウンドクロールを使用
        """)

    with st.expander("### 📝 表記ゆれ・禁止表現"):
        st.markdown("""
各ページのテキストを **toyota_lexus用語リスト** と照合し、誤表記・禁止表現を検出します。

**辞書について:**
- デフォルト: リポジトリ内の `hyoki_dict.xlsx`（167件）
- 更新: サイドバーの「辞書を更新」から Excel をアップロード → 差分確認 → 適用
- 永続化: 適用後に表示されるダウンロードボタンで取得し、`hyoki_dict.xlsx` としてリポジトリに commit

**カスタムエントリ:** サイドバーのテキストエリアに `NG表記|推奨表記` 形式で追加可能（正規表現も使用可）

**出力:**
- ページ単位サマリ（表記ゆれ件数 / 禁止表現件数）
- 修正指示書（1指摘1行。「修正済み」列に ✅ を入れて管理推奨）
        """)

    with st.expander("### 🔍 アプリ・機能検出"):
        st.markdown("""
ページの HTML を静的解析し、**WebアプリやUIコンポーネント** の有無を検出します。

**検出カテゴリ:**

| カテゴリ | 検出対象の例 |
|---------|------------|
| JSフレームワーク | Next.js, React, Vue.js, Angular, Nuxt.js |
| 認証 | ログインページ, マイページ, トヨタID |
| フォーム | 問い合わせ, 試乗予約, 申込, 検索 |
| 設定ツール | グレード選択, 見積もりシミュレーター, 販売店検索 |
| メディア | Brightcove動画, YouTube, HTML5 video, Google Maps, 360°ビュー |
| SPA/PWA | PWAマニフェスト, SPAルーティング, Fetch/XHR多用 |
| 外部サービス | チャットボット, SNS共有, GTM/Adobe Analytics |
| トヨタ固有 | T-Connect, KINTO, ファイナンス, カーラインアップ |

**アプリ性スコア:** 検出されたカテゴリ数。スコア3以上は複雑なインタラクションがあるページとして要注目。
        """)

    with st.expander("### 🕷️ バックグラウンドクロール"):
        st.markdown("""
**ブラウザを閉じても継続する** バックグラウンドジョブとして大規模クロールを実行します。

**通常モードとの違い:**

| 項目 | 通常モード | BGクロール |
|------|-----------|-----------|
| 最大ページ数 | 〜数十ページ推奨 | 最大 1,000 ページ |
| ブラウザ常駐 | 必要 | 不要 |
| 複数チェック同時実行 | 不可 | 可能 |
| URLソース | Excel / 直接入力 / クロール | クロール or コンテンツ管理票Excel |

**使い方:**
1. サイドバーでジョブ設定（URLソース・実行するチェック）を設定
2. 「▶ チェック実行」でジョブ開始 → **ジョブID** が発行される
3. 「🔍 ジョブ確認」タブにIDを入力し「🔄 更新」で進捗確認
4. `done` になったら結果を確認・ダウンロード

**スタック検知:** 最終更新から10分以上経過すると警告が表示されます。

**コンテンツ管理票 Excel の対応シート:**

| シート名 | URL列 | データ開始行 |
|---------|-------|------------|
| 運用サイトマップ | H列（フルURL） | 6行目 |
| コンテンツ_車種 | C列（相対パス） | 5行目 |
| コンテンツ_車種横断 | C列（相対パス） | 3行目 |
| TQP非検知info・informationフォルダ | B列（フルURL） | 2行目 |
| CMS移行 | A列（フルURL） | 2行目（対応列が「移行」を含む行のみ） |

相対パス（`/info/...` 等）は自動で `https://toyota.jp` が補完されます。
シート横断で重複URLは1件に集約されます。
        """)

    with st.expander("### ❓ よくある質問・注意事項"):
        st.markdown("""
**Q. リンクチェックで「⚠️ 確認不可」が大量に出る**
> チェックサーバー（Streamlit Cloud）の IP が toyota.jp にブロックされている可能性があります。
> 実際にはリンク切れでないケースが多いです。ブラウザで個別確認してください。

**Q. クロールが遅い / 途中で止まる**
> toyota.jp はリクエスト間隔を設けています。大量ページは BGクロールを使用してください。

**Q. 表記ゆれ辞書をアップロードしたが再起動後に消えた**
> `/tmp` はアプリ再起動で消去されます。辞書を永続化するには：
> ① 差分を適用 → ② 表示されるダウンロードボタンで `hyoki_dict.xlsx` を保存
> → ③ リポジトリに commit & push

**Q. Core Web Vitals が「N/A」になる**
> PSI APIが対象URLを計測できない場合（ログイン必須ページ等）に発生します。

**Q. Excel の「運用サイトマップ」シートが見つからないエラー**
> シート名が完全一致している必要があります。シート名を確認してください。

**Q. BGジョブIDを忘れた**
> 「🔍 ジョブ確認」タブの「📋 ジョブ履歴から選択」ドロップダウンに直近20件が表示されます。
        """)

    st.stop()

# ─── 通常モード: ボタンを押すまで待機 ──────────────────────────────
if not run_btn:
    st.info("サイドバーで設定を入力し「▶ チェック実行」を押してください。")
    st.stop()

# ─── URL 収集（通常モード） ──────────────────────────────────────────
if url_source == "Excelファイルをアップロード":
    if not uploaded_file:
        st.warning("Excelファイルをアップロードしてください。")
        st.stop()
    if not excel_sheets:
        st.warning("読み込むシートを1つ以上選択してください。")
        st.stop()
    with st.spinner("Excelを読み込み中..."):
        urls = load_urls_from_excel(uploaded_file, excel_sheets)
    if limit:
        urls = urls[:limit]

elif url_source == "URLを直接入力":
    urls = manual_urls

else:  # 🕷️ クロール
    from urllib.parse import urlparse
    from collections import deque
    from fetcher import extract_links as _extract_links

    if not crawl_start or not crawl_start.startswith("http"):
        st.error("クロール開始URLを入力してください。")
        st.stop()

    base_domain = urlparse(crawl_start).netloc
    visited: set = set()
    queue: deque = deque([(crawl_start.rstrip("/"), 0)])
    urls: list = []

    _path_label = f"、パス: `{crawl_path_filter}`" if crawl_path_filter else ""
    st.info(f"**{crawl_start}** からクロール中（最大 {crawl_max} ページ、深さ {crawl_depth}{_path_label}）")
    crawl_progress = st.progress(0, text="クロール中...")
    crawl_status = st.empty()

    while queue and len(urls) < crawl_max:
        url, depth = queue.popleft()
        norm = url.rstrip("/").split("?")[0].split("#")[0]
        if norm in visited:
            continue
        visited.add(norm)

        code, html = fetch_html(url)
        if code == 200 and html:
            _url_path = urlparse(url).path
            if not crawl_path_filter or _url_path.startswith(crawl_path_filter):
                urls.append(url)
                crawl_progress.progress(
                    min(len(urls) / crawl_max, 1.0),
                    text=f"収集中... {len(urls)} / {crawl_max} ページ",
                )
                crawl_status.caption(f"→ {url}")

            if depth < crawl_depth:
                for link in _extract_links(html, url):
                    _lp = urlparse(link)
                    if _lp.netloc == base_domain:
                        if not crawl_path_filter or _lp.path.startswith(crawl_path_filter):
                            clean = link.rstrip("/").split("?")[0].split("#")[0]
                            if clean not in visited:
                                queue.append((link, depth + 1))

        time.sleep(REQUEST_DELAY)

    crawl_progress.empty()
    crawl_status.empty()
    st.success(f"クロール完了: **{len(urls)} ページ**を収集しました")

if not urls:
    st.error("URLが見つかりません。")
    st.stop()

st.success(f"対象URL: **{len(urls)} 件**")

# ─── 各チェック実行 ───────────────────────────────────

# ════════════════════════════════════════
# 📄 タイトル取得
# ════════════════════════════════════════
if check_type == "📄 タイトル取得":
    from concurrent.futures import ThreadPoolExecutor, as_completed

    results = []
    progress = st.progress(0, text="取得中...")
    status_box = st.empty()

    def _fetch(url):
        code, html = fetch_html(url)
        if code == 200 and html:
            meta = extract_meta(html)
            return {"URL": url, "ステータス": code,
                    "タイトル": meta["short_title"],
                    "description": meta["description"][:100],
                    "判定": "✅" if meta["short_title"] else "⚠️ titleなし"}
        return {"URL": url, "ステータス": code, "タイトル": "", "description": "", "判定": "❌"}

    with ThreadPoolExecutor(max_workers=CONCURRENT_WORKERS) as ex:
        futures = {ex.submit(_fetch, u): u for u in urls}
        done = 0
        for future in as_completed(futures):
            results.append(future.result())
            done += 1
            progress.progress(done / len(urls), text=f"取得中... {done}/{len(urls)}")
            time.sleep(REQUEST_DELAY)

    progress.empty()
    df = pd.DataFrame(results)
    ng = df[df["判定"] != "✅"]

    col1, col2 = st.columns(2)
    col1.metric("✅ 取得成功", len(df[df["判定"] == "✅"]))
    col2.metric("⚠️ 要確認", len(ng))

    st.subheader("結果一覧")
    st.dataframe(df, use_container_width=True, height=500)

    st.download_button("📥 Excelダウンロード", to_excel_bytes(df),
                       "titles_result.xlsx", use_container_width=True)

# ════════════════════════════════════════
# 🔗 リンクチェック
# ════════════════════════════════════════
elif check_type == "🔗 リンクチェック":
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from fetcher import extract_resources, extract_css_urls
    from urllib.parse import urlparse

    # Phase 1: 各ページのリソースを収集
    st.caption("Phase 1/2 — ページ内リソースを収集中（リンク・CSS・JS・画像・メディア・CSSリソース）")
    progress1 = st.progress(0, text="収集中...")
    page_resources: dict = {}  # {src_url: {"status": int, "resources": [(url, type), ...]}}

    _css_res_types = {"画像(CSS)", "フォント(CSS)", "CSSリソース"}

    def _collect(url):
        code, html = fetch_html(url)
        resources = extract_resources(html, url) if (code == 200 and html) else []

        # CSS内リソースが選択対象の場合のみCSSファイルをフェッチ
        if not selected_res_types or selected_res_types & _css_res_types:
            for css_url in [u for u, t in resources if t == "CSS/スタイル"]:
                css_code, css_text = fetch_html(css_url)
                if css_code == 200 and css_text:
                    resources.extend(extract_css_urls(css_text, css_url))

        if selected_res_types:
            resources = [(u, t) for u, t in resources if t in selected_res_types]
        if toyota_only:
            resources = [(u, t) for u, t in resources if urlparse(u).netloc == "toyota.jp"]
        return url, code, resources

    with ThreadPoolExecutor(max_workers=CONCURRENT_WORKERS) as ex:
        futs = {ex.submit(_collect, u): u for u in urls}
        done = 0
        for f in as_completed(futs):
            src, code, resources = f.result()
            page_resources[src] = {"status": code, "resources": resources}
            done += 1
            progress1.progress(done / len(urls), text=f"リソース収集中... {done}/{len(urls)}")
            time.sleep(REQUEST_DELAY)
    progress1.empty()

    all_resource_urls = list({u for d in page_resources.values() for u, _ in d["resources"]})
    st.caption(f"Phase 2/2 — 発見リソース {len(all_resource_urls)} 件のステータスを確認中")

    # Phase 2: 収集したリソースのステータスチェック
    progress2 = st.progress(0, text="チェック中...")
    resource_status: dict = {}

    def _ping(url):
        code, _ = fetch_html(url)
        return url, code

    with ThreadPoolExecutor(max_workers=CONCURRENT_WORKERS) as ex:
        futs = {ex.submit(_ping, u): u for u in all_resource_urls}
        done = 0
        for f in as_completed(futs):
            url, code = f.result()
            resource_status[url] = code
            done += 1
            progress2.progress(done / len(all_resource_urls), text=f"確認中... {done}/{len(all_resource_urls)}")
            time.sleep(REQUEST_DELAY)
    progress2.empty()

    # Phase 3: 発見ページ × リソースURL で集計
    broken_rows = []      # 404 / 410 — 確実なリンク切れ
    unverif_rows = []     # status=0  — 接続失敗（切れているとは断言できない）
    all_rows = []
    for src in sorted(page_resources):
        d = page_resources[src]
        for res_url, res_type in d["resources"]:
            code = resource_status.get(res_url, 0)
            if code in (404, 410):
                judgment = "❌ 切れ"
            elif code == 0:
                judgment = "⚠️ 確認不可"
            else:
                judgment = "✅ 正常"
            row = {
                "発見ページ": src,
                "リソースURL": res_url,
                "種別": res_type,
                "ドメイン": urlparse(res_url).netloc,
                "ステータス": code,
                "判定": judgment,
            }
            all_rows.append(row)
            if judgment == "❌ 切れ":
                broken_rows.append(row)
            elif judgment == "⚠️ 確認不可":
                unverif_rows.append(row)

    broken_pages = len({r["発見ページ"] for r in broken_rows})
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("対象ページ数", len(urls))
    col2.metric("❌ リンク切れ件数", len(broken_rows))
    col3.metric("❌ リンク切れページ", broken_pages)
    col4.metric("⚠️ 確認不可", len(unverif_rows))

    if broken_rows:
        st.error(f"リンク切れ: {len(broken_rows)} 件（{broken_pages} ページで発見）")
        st.subheader("❌ リンク切れ一覧（404 / 410）")
        df_broken = pd.DataFrame(broken_rows)
        types = ["すべて"] + sorted(df_broken["種別"].unique().tolist())
        sel = st.selectbox("種別フィルタ", types)
        disp = df_broken if sel == "すべて" else df_broken[df_broken["種別"] == sel]
        st.dataframe(disp, use_container_width=True, height=400)
        c1, c2 = st.columns(2)
        c1.download_button("📥 Excel", to_excel_bytes(df_broken),
                           "links_broken.xlsx", use_container_width=True, key="dl_broken_xlsx")
        c2.download_button("📥 CSV", df_broken.to_csv(index=False).encode("utf-8-sig"),
                           "links_broken.csv", mime="text/csv",
                           use_container_width=True, key="dl_broken_csv")
    else:
        st.success("リンク切れは検出されませんでした ✅")

    if unverif_rows:
        with st.expander(f"⚠️ 確認不可 {len(unverif_rows)} 件（接続できなかったリソース）"):
            st.caption(
                "チェックサーバーから到達できなかったリソースです（status=0）。"
                "アクセス制限・レートリミット・ネットワーク経路の問題が原因の場合が多く、"
                "実際にはリンク切れでないケースがあります。ブラウザで個別に確認してください。"
            )
            df_unv = pd.DataFrame(unverif_rows)
            st.dataframe(df_unv, use_container_width=True, height=300)
            c1u, c2u = st.columns(2)
            c1u.download_button("📥 Excel", to_excel_bytes(df_unv),
                                "links_unverifiable.xlsx", use_container_width=True,
                                key="dl_unv_xlsx")
            c2u.download_button("📥 CSV", df_unv.to_csv(index=False).encode("utf-8-sig"),
                                "links_unverifiable.csv", mime="text/csv",
                                use_container_width=True, key="dl_unv_csv")

    with st.expander(f"全リソース一覧（{len(all_rows)} 件）"):
        df_all = pd.DataFrame(all_rows)
        st.dataframe(df_all, use_container_width=True, height=400)
        ca, cb = st.columns(2)
        ca.download_button("📥 全件Excel", to_excel_bytes(df_all),
                           "links_all.xlsx", use_container_width=True, key="dl_all_xlsx")
        cb.download_button("📥 全件CSV", df_all.to_csv(index=False).encode("utf-8-sig"),
                           "links_all.csv", mime="text/csv",
                           use_container_width=True, key="dl_all_csv")

# ════════════════════════════════════════
# 🌐 外部リンクチェック
# ════════════════════════════════════════
elif check_type == "🌐 外部リンクチェック":
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from fetcher import extract_links as _ext_extract_links
    from urllib.parse import urlparse as _ext_urlparse

    base_domain = _ext_urlparse(urls[0]).netloc if urls else "toyota.jp"

    # Phase 1: 各ページから外部リンクを収集
    st.caption("Phase 1/2 — ページ内の外部リンクを収集中")
    progress1 = st.progress(0, text="収集中...")
    page_ext_links: dict = {}

    def _collect_ext(url):
        code, html = fetch_html(url)
        if code != 200 or not html:
            return url, []
        all_links = _ext_extract_links(html, url)
        return url, [lk for lk in all_links
                     if _ext_urlparse(lk).netloc not in (base_domain, "")]

    with ThreadPoolExecutor(max_workers=CONCURRENT_WORKERS) as ex:
        futs = {ex.submit(_collect_ext, u): u for u in urls}
        done = 0
        for f in as_completed(futs):
            pg, links = f.result()
            page_ext_links[pg] = links
            done += 1
            progress1.progress(done / len(urls), text=f"収集中... {done}/{len(urls)}")
            time.sleep(REQUEST_DELAY)
    progress1.empty()

    all_ext_urls = list({lk for links in page_ext_links.values() for lk in links})

    if not all_ext_urls:
        st.info("外部リンクは検出されませんでした。")
        st.stop()

    st.caption(f"Phase 2/2 — 外部リンク {len(all_ext_urls)} 件のステータスを確認中")
    progress2 = st.progress(0, text="確認中...")
    ext_status: dict = {}

    def _ping_ext(url):
        code, _ = fetch_html(url)
        return url, code

    with ThreadPoolExecutor(max_workers=CONCURRENT_WORKERS) as ex:
        futs = {ex.submit(_ping_ext, u): u for u in all_ext_urls}
        done = 0
        for f in as_completed(futs):
            url, code = f.result()
            ext_status[url] = code
            done += 1
            progress2.progress(done / len(all_ext_urls), text=f"確認中... {done}/{len(all_ext_urls)}")
            time.sleep(REQUEST_DELAY)
    progress2.empty()

    # Phase 3: 集計
    ext_broken, ext_restricted, ext_unverif, ext_all = [], [], [], []
    for src in sorted(page_ext_links):
        for ext_url in page_ext_links[src]:
            code = ext_status.get(ext_url, 0)
            domain = _ext_urlparse(ext_url).netloc
            if code in (404, 410):
                judgment = "❌ 切れ"
            elif code == 403:
                judgment = "🔒 アクセス制限"
            elif code == 0:
                judgment = "⚠️ 確認不可"
            else:
                judgment = "✅ 正常"
            row = {
                "発見ページ": src, "外部リンクURL": ext_url,
                "ドメイン": domain, "ステータス": code, "判定": judgment,
            }
            ext_all.append(row)
            if judgment == "❌ 切れ":
                ext_broken.append(row)
            elif judgment == "🔒 アクセス制限":
                ext_restricted.append(row)
            elif judgment == "⚠️ 確認不可":
                ext_unverif.append(row)

    domain_counts = {}
    for r in ext_all:
        d = r["ドメイン"]
        domain_counts[d] = domain_counts.get(d, 0) + 1

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("外部リンク総数", len(ext_all))
    c2.metric("❌ 切れ", len(ext_broken))
    c3.metric("🔒 アクセス制限", len(ext_restricted))
    c4.metric("⚠️ 確認不可", len(ext_unverif))

    if ext_broken:
        st.error(f"外部リンク切れ: {len(ext_broken)} 件")
        st.subheader("❌ 切れた外部リンク（404 / 410）")
        st.dataframe(pd.DataFrame(ext_broken), use_container_width=True, height=300)
        c1x, c2x = st.columns(2)
        c1x.download_button("📥 Excel", to_excel_bytes(pd.DataFrame(ext_broken)),
                            "ext_broken.xlsx", use_container_width=True, key="dl_extb_x")
        c2x.download_button("📥 CSV", pd.DataFrame(ext_broken).to_csv(index=False).encode("utf-8-sig"),
                            "ext_broken.csv", mime="text/csv",
                            use_container_width=True, key="dl_extb_c")
    else:
        st.success("外部リンク切れは検出されませんでした ✅")

    if ext_restricted:
        with st.expander(f"🔒 アクセス制限 {len(ext_restricted)} 件（403）"):
            st.caption("チェックサーバーからのアクセスが拒否されましたが、ブラウザでは閲覧できる場合があります。")
            st.dataframe(pd.DataFrame(ext_restricted), use_container_width=True, height=200)

    if ext_unverif:
        with st.expander(f"⚠️ 確認不可 {len(ext_unverif)} 件（接続失敗）"):
            st.caption("タイムアウト・接続拒否等。実際には正常なリンクの可能性があります。")
            st.dataframe(pd.DataFrame(ext_unverif), use_container_width=True, height=200)

    st.subheader("ドメイン別集計")
    df_dom = pd.DataFrame(
        [{"ドメイン": d, "リンク数": cnt} for d, cnt in
         sorted(domain_counts.items(), key=lambda x: -x[1])]
    )
    st.dataframe(df_dom, use_container_width=True, height=300)

    with st.expander(f"全外部リンク一覧 ({len(ext_all)} 件)"):
        df_ext_all = pd.DataFrame(ext_all)
        _sel_d = st.selectbox("ドメインフィルタ",
                              ["すべて"] + sorted(domain_counts.keys()),
                              key="ext_dom_filter")
        disp_ext = df_ext_all if _sel_d == "すべて" else df_ext_all[df_ext_all["ドメイン"] == _sel_d]
        st.dataframe(disp_ext, use_container_width=True, height=350)
        ca, cb = st.columns(2)
        ca.download_button("📥 全件Excel", to_excel_bytes(df_ext_all),
                           "extlinks_all.xlsx", use_container_width=True, key="dl_extall_x")
        cb.download_button("📥 全件CSV", df_ext_all.to_csv(index=False).encode("utf-8-sig"),
                           "extlinks_all.csv", mime="text/csv",
                           use_container_width=True, key="dl_extall_c")

# ════════════════════════════════════════
# ⚡ Core Web Vitals
# ════════════════════════════════════════
elif check_type == "⚡ Core Web Vitals":
    import json, urllib.request, urllib.parse
    from config import PSI_ENDPOINT, THRESHOLDS

    results = []
    progress = st.progress(0, text="計測中...")

    def _cwv(url):
        params = {"url": url, "strategy": strategy, "category": "performance"}
        if psi_key:
            params["key"] = psi_key
        api_url = f"{PSI_ENDPOINT}?{urllib.parse.urlencode(params)}"
        try:
            req = urllib.request.Request(api_url, headers={"User-Agent": "toyota-check/1.0"})
            with urllib.request.urlopen(req, timeout=30) as res:
                data = json.loads(res.read().decode())
        except Exception as e:
            return {"URL": url, "エラー": str(e)}

        audits = data.get("lighthouseResult", {}).get("audits", {})
        score = data.get("lighthouseResult", {}).get("categories", {}).get("performance", {}).get("score")

        def ms(k): v = audits.get(k, {}).get("numericValue"); return round(v) if v else None
        lcp = ms("largest-contentful-paint")
        cls = round(audits.get("cumulative-layout-shift", {}).get("numericValue", 0), 3)
        inp = ms("interaction-to-next-paint") or ms("total-blocking-time")

        def rate(v, m):
            t = THRESHOLDS.get(m, {})
            if v is None: return "N/A"
            return "GOOD" if v <= t["good"] else ("NI" if v <= t["poor"] else "POOR")

        ng = [m for m, v in [("LCP", lcp), ("CLS", cls), ("INP", inp)]
              if rate(v, m) in ("NI", "POOR")]
        return {
            "URL": url, "スコア": round(score * 100) if score else None,
            "LCP(ms)": lcp, "LCP判定": rate(lcp, "LCP"),
            "CLS": cls,     "CLS判定": rate(cls, "CLS"),
            "INP(ms)": inp, "INP判定": rate(inp, "INP"),
            "KPI未達": ", ".join(ng) if ng else "✅",
        }

    for i, url in enumerate(urls):
        results.append(_cwv(url))
        progress.progress((i + 1) / len(urls), text=f"計測中... {i+1}/{len(urls)}")
        time.sleep(1.0)

    progress.empty()
    df = pd.DataFrame(results)
    ng_df = df[df["KPI未達"] != "✅"] if "KPI未達" in df.columns else pd.DataFrame()

    col1, col2 = st.columns(2)
    col1.metric("✅ KPI達成", len(df) - len(ng_df))
    col2.metric("❌ KPI未達", len(ng_df))

    if len(ng_df):
        st.subheader("❌ KPI未達ページ")
        st.dataframe(ng_df, use_container_width=True)

    with st.expander("全件表示"):
        st.dataframe(df, use_container_width=True, height=400)

    st.download_button("📥 Excelダウンロード", to_excel_bytes(df),
                       "cwv_result.xlsx", use_container_width=True)

# ════════════════════════════════════════
# 📝 表記ゆれ・禁止表現
# ════════════════════════════════════════
elif check_type == "📝 表記ゆれ・禁止表現":
    import re
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from fetcher import extract_meta, extract_text_blocks
    from config import KINSHI_PATTERNS
    from dict_loader import load_for_check

    # Excel辞書 + カスタム追加エントリをマージ
    merged_dict = load_for_check()
    for line in custom_dict_raw.splitlines():
        line = line.strip()
        if "|" in line:
            wrong, correct = line.split("|", 1)
            merged_dict[re.escape(wrong.strip())] = correct.strip()

    results = []
    progress = st.progress(0, text="解析中...")

    def _content(url):
        code, html = fetch_html(url)
        if code != 200 or not html:
            return {"url": url, "code": code, "title": "", "findings": [], "判定": "❌ 取得失敗"}
        meta = extract_meta(html)
        text = " ".join(extract_text_blocks(html))
        findings = []

        for pattern, recommended in merged_dict.items():
            for match in sorted(set(re.findall(pattern, text, re.I))):
                findings.append({
                    "URL": url,
                    "ページタイトル": meta["short_title"],
                    "種別": "表記ゆれ",
                    "発見テキスト": match,
                    "推奨表記": recommended,
                    "修正済み": "",
                })

        for pattern, reason in KINSHI_PATTERNS:
            for match in sorted(set(re.findall(pattern, text))):
                findings.append({
                    "URL": url,
                    "ページタイトル": meta["short_title"],
                    "種別": "禁止表現",
                    "発見テキスト": match,
                    "推奨表記": f"【要確認】{reason}",
                    "修正済み": "",
                })

        return {
            "url": url, "code": code, "title": meta["short_title"],
            "findings": findings,
            "判定": "⚠️ 要確認" if findings else "✅",
        }

    with ThreadPoolExecutor(max_workers=CONCURRENT_WORKERS) as ex:
        futures = {ex.submit(_content, u): u for u in urls}
        done = 0
        for future in as_completed(futures):
            results.append(future.result())
            done += 1
            progress.progress(done / len(urls), text=f"解析中... {done}/{len(urls)}")
            time.sleep(REQUEST_DELAY)

    progress.empty()

    # ページ単位サマリ
    summary_rows = [{
        "URL": r["url"],
        "ページタイトル": r["title"],
        "ステータス": r["code"],
        "表記ゆれ件数": sum(1 for f in r["findings"] if f["種別"] == "表記ゆれ"),
        "禁止表現件数": sum(1 for f in r["findings"] if f["種別"] == "禁止表現"),
        "判定": r["判定"],
    } for r in results]

    # 指示書（1指摘1行）
    all_findings = sorted(
        [f for r in results for f in r["findings"]],
        key=lambda x: (x["種別"], x["URL"])
    )
    instruction_rows = [{"No.": i, **f} for i, f in enumerate(all_findings, 1)]

    df_summary = pd.DataFrame(summary_rows)
    ng_count = len([r for r in results if r["判定"] == "⚠️ 要確認"])

    col1, col2, col3 = st.columns(3)
    col1.metric("✅ 問題なし", len(results) - ng_count)
    col2.metric("⚠️ 要確認ページ", ng_count)
    col3.metric("📋 指摘件数合計", len(instruction_rows))

    st.subheader("ページ単位サマリ")
    st.dataframe(df_summary, use_container_width=True, height=350)

    if instruction_rows:
        st.subheader("📋 修正指示書")
        df_inst = pd.DataFrame(instruction_rows)

        kinds = ["すべて"] + sorted(df_inst["種別"].unique().tolist())
        sel_kind = st.selectbox("種別フィルタ", kinds)
        disp_inst = df_inst if sel_kind == "すべて" else df_inst[df_inst["種別"] == sel_kind]
        st.dataframe(disp_inst, use_container_width=True, height=420)

        c1, c2 = st.columns(2)
        c1.download_button(
            "📥 指示書 Excel",
            to_excel_bytes(df_inst),
            "修正指示書.xlsx",
            use_container_width=True,
        )
        c2.download_button(
            "📥 指示書 CSV",
            df_inst.to_csv(index=False).encode("utf-8-sig"),
            "修正指示書.csv",
            mime="text/csv",
            use_container_width=True,
            key="dl_inst_csv",
        )
    else:
        st.success("問題のある表現は検出されませんでした ✅")

# ════════════════════════════════════════
# 🔍 アプリ・機能検出
# ════════════════════════════════════════
elif check_type == "🔍 アプリ・機能検出":
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from fetcher import extract_meta
    from detector import summarize, CATEGORIES

    rows = []
    progress = st.progress(0, text="解析中...")

    def _detect(url):
        code, html = fetch_html(url)
        if code != 200 or not html:
            return {"URL": url, "タイトル": "", "アプリ性スコア": 0, "検出機能": f"❌ HTTP {code}",
                    **{c: "" for c in CATEGORIES}}
        meta = extract_meta(html)
        return summarize(url, html, meta["short_title"])

    with ThreadPoolExecutor(max_workers=CONCURRENT_WORKERS) as ex:
        futs = {ex.submit(_detect, u): u for u in urls}
        done = 0
        for f in as_completed(futs):
            rows.append(f.result())
            done += 1
            progress.progress(done / len(urls), text=f"解析中... {done}/{len(urls)}")
            time.sleep(REQUEST_DELAY)

    progress.empty()
    df_app = pd.DataFrame(rows)

    # ── サマリーメトリクス ────────────────────────────────────────────
    n_with_feat = len(df_app[df_app["アプリ性スコア"] > 0])
    n_high      = len(df_app[df_app["アプリ性スコア"] >= 3])
    cat_counts  = {c: int((df_app[c] != "").sum()) for c in CATEGORIES}

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("解析ページ数", len(df_app))
    c2.metric("機能検出あり", n_with_feat)
    c3.metric("スコア3以上（高複雑度）", n_high)
    c4.metric("最高スコア", int(df_app["アプリ性スコア"].max()) if len(df_app) else 0)

    # ── カテゴリ別ヒット数 ─────────────────────────────────────────────
    st.subheader("カテゴリ別検出ページ数")
    df_cat = pd.DataFrame(
        [{"カテゴリ": c, "検出ページ数": cat_counts[c]} for c in CATEGORIES if cat_counts[c] > 0]
    ).sort_values("検出ページ数", ascending=False)
    st.bar_chart(df_cat.set_index("カテゴリ")["検出ページ数"])

    # ── フィルター付きページ一覧 ───────────────────────────────────────
    st.subheader("ページ一覧（アプリ性スコア順）")
    _cat_opts = ["すべて"] + [c for c in CATEGORIES if cat_counts.get(c, 0) > 0]
    _cat_sel = st.selectbox("カテゴリでフィルタ", _cat_opts)
    _score_min = st.slider("アプリ性スコア（最小）", 0, len(CATEGORIES), 1)

    df_disp = df_app.copy()
    if _cat_sel != "すべて":
        df_disp = df_disp[df_disp[_cat_sel] != ""]
    df_disp = df_disp[df_disp["アプリ性スコア"] >= _score_min]
    df_disp = df_disp.sort_values("アプリ性スコア", ascending=False)

    # 表示用: 全カラムを表示するとワイドになるので要約カラムを優先
    _display_cols = ["URL", "タイトル", "アプリ性スコア", "検出機能"]
    st.dataframe(df_disp[_display_cols], use_container_width=True, height=460)

    # ── カテゴリ別ピボット（Excel向け） ────────────────────────────────
    with st.expander("カテゴリ別詳細（ピボット表示）"):
        _pivot_cols = ["URL", "タイトル", "アプリ性スコア"] + CATEGORIES
        st.dataframe(
            df_app.sort_values("アプリ性スコア", ascending=False)[_pivot_cols],
            use_container_width=True,
            height=400,
        )

    # ── ダウンロード ───────────────────────────────────────────────────
    _all_cols = ["URL", "タイトル", "アプリ性スコア", "検出機能"] + CATEGORIES
    df_dl = df_app.sort_values("アプリ性スコア", ascending=False)[_all_cols]
    cx, cy = st.columns(2)
    cx.download_button(
        "📥 Excel（全件）",
        to_excel_bytes(df_dl),
        "app_detection.xlsx",
        use_container_width=True,
        key="dl_app_xlsx",
    )
    cy.download_button(
        "📥 CSV（全件）",
        df_dl.to_csv(index=False).encode("utf-8-sig"),
        "app_detection.csv",
        mime="text/csv",
        use_container_width=True,
        key="dl_app_csv",
    )

# ════════════════════════════════════════
# 📌 リンク元調査
# ════════════════════════════════════════
elif check_type == "📌 リンク元調査":
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from fetcher import extract_links as _bl_extract_links
    from fetcher import extract_meta as _bl_extract_meta

    if not backlink_query:
        st.warning("サイドバーで「調査対象URL」を入力してください。")
        st.stop()

    st.info(
        f"**{len(urls)} ページ**を対象に「`{backlink_query}`」へのリンクを探しています。\n\n"
        "内部リンク・外部リンクの両方を対象とします。"
    )
    progress = st.progress(0, text="解析中...")

    bl_found: list = []

    def _bl_check(url):
        code, html = fetch_html(url)
        if code != 200 or not html:
            return []
        links = _bl_extract_links(html, url)
        matched = [lk for lk in links if backlink_query in lk]
        if not matched:
            return []
        meta = _bl_extract_meta(html)
        return [
            {"発見ページURL": url, "ページタイトル": meta["short_title"], "マッチしたリンク": lk}
            for lk in dict.fromkeys(matched)  # 同一ページ内の重複リンクは1件に集約
        ]

    with ThreadPoolExecutor(max_workers=CONCURRENT_WORKERS) as ex:
        futs = {ex.submit(_bl_check, u): u for u in urls}
        done = 0
        for f in as_completed(futs):
            bl_found.extend(f.result())
            done += 1
            progress.progress(done / len(urls), text=f"解析中... {done}/{len(urls)}")
            time.sleep(REQUEST_DELAY)

    progress.empty()

    pages_with_link = list(dict.fromkeys(r["発見ページURL"] for r in bl_found))

    c1, c2, c3 = st.columns(3)
    c1.metric("調査ページ数", len(urls))
    c2.metric("📌 リンク元ページ数", len(pages_with_link))
    c3.metric("マッチリンク総数", len(bl_found))

    if bl_found:
        df_bl = pd.DataFrame(bl_found)

        st.subheader(f"「{backlink_query}」へのリンクを含むページ一覧")

        # ページ単位サマリ（1ページ1行）
        df_bl_summary = (
            df_bl.groupby(["発見ページURL", "ページタイトル"])["マッチしたリンク"]
            .count().reset_index()
            .rename(columns={"マッチしたリンク": "リンク数"})
            .sort_values("リンク数", ascending=False)
        )
        st.dataframe(df_bl_summary, use_container_width=True, height=350)

        with st.expander(f"マッチしたリンク詳細（{len(bl_found)} 件）"):
            st.dataframe(df_bl, use_container_width=True, height=350)

        c1x, c2x = st.columns(2)
        c1x.download_button(
            "📥 Excel（サマリ）",
            to_excel_bytes(df_bl_summary),
            "backlinks_summary.xlsx",
            use_container_width=True,
            key="dl_bl_sum_x",
        )
        c2x.download_button(
            "📥 Excel（詳細）",
            to_excel_bytes(df_bl),
            "backlinks_detail.xlsx",
            use_container_width=True,
            key="dl_bl_det_x",
        )
        st.download_button(
            "📥 CSV（詳細）",
            df_bl.to_csv(index=False).encode("utf-8-sig"),
            "backlinks_detail.csv",
            mime="text/csv",
            use_container_width=True,
            key="dl_bl_det_c",
        )
    else:
        st.success(f"「{backlink_query}」へのリンクを含むページは見つかりませんでした。")
