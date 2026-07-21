"""サイトマップ管理モード

tjpコンテンツ管理表.xlsx（運用サイトマップ）をアプリ上で閲覧・編集・保守する。
  - 第6階層までの階層列を自動生成し、表示階層を指定してビューを切替
  - URL一覧（編集可・行追加可）／ディレクトリ集計の2ビュー
  - CSV / Excel ダウンロード
  - HTML（静的）かアプリケーション（動的）かをURLルールで自動判別
  - サイトをクロールして一覧に無いURL・より深い階層を検知し、一覧へ追加
"""

from __future__ import annotations

import io
import json
import re
import time
from collections import deque
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlparse, urljoin, urldefrag

import pandas as pd

# ─── 永続ストレージ ──────────────────────────────────────────
STORE_DIR = Path(__file__).parent / "data"
STORE_CSV = STORE_DIR / "sitemap_master.csv"
META_JSON = STORE_DIR / "sitemap_meta.json"
GH_PATH = "data/sitemap_master.csv"


def _gh_cfg():
    """GitHubバックアップ設定（st.secrets に GITHUB_TOKEN があれば有効）。"""
    try:
        import streamlit as st
        tok = st.secrets.get("GITHUB_TOKEN", "")
        repo = st.secrets.get("GITHUB_REPO", "mutulab/toyota-check")
        return (tok, repo) if tok else None
    except Exception:
        return None


def load_store() -> tuple[pd.DataFrame | None, dict]:
    """保存済みマスタを読み込む（ローカル → 無ければGitHubから復元）。"""
    meta: dict = {}
    if META_JSON.exists():
        try:
            meta = json.loads(META_JSON.read_text(encoding="utf-8"))
        except Exception:
            meta = {}
    if STORE_CSV.exists():
        return pd.read_csv(STORE_CSV, dtype="string").fillna(pd.NA), meta

    gh = _gh_cfg()
    if gh:
        import requests
        tok, repo = gh
        try:
            r = requests.get(
                f"https://api.github.com/repos/{repo}/contents/{GH_PATH}",
                headers={"Authorization": f"Bearer {tok}",
                         "Accept": "application/vnd.github.raw"},
                timeout=20)
            if r.status_code == 200:
                STORE_DIR.mkdir(exist_ok=True)
                STORE_CSV.write_bytes(r.content)
                meta["restored_from_github"] = True
                return pd.read_csv(STORE_CSV, dtype="string").fillna(pd.NA), meta
        except Exception:
            pass
    return None, meta


def _gh_push(action: str) -> tuple[bool, str]:
    """マスタCSVをGitHubへバックアップ。(成功可否, 失敗理由) を返す。"""
    gh = _gh_cfg()
    if not gh:
        return False, "GITHUB_TOKEN未設定"
    import base64
    import requests
    tok, repo = gh
    url = f"https://api.github.com/repos/{repo}/contents/{GH_PATH}"
    hdr = {"Authorization": f"Bearer {tok}",
           "Accept": "application/vnd.github+json"}
    try:
        sha = None
        r = requests.get(url, headers=hdr, timeout=20)
        if r.status_code == 200:
            sha = r.json().get("sha")
        elif r.status_code not in (404,):
            return False, f"GitHub照会エラー HTTP {r.status_code}: {r.json().get('message', '')}"
        body = {"message": f"data: サイトマップマスタ更新（{action}）",
                "content": base64.b64encode(STORE_CSV.read_bytes()).decode()}
        if sha:
            body["sha"] = sha
        r = requests.put(url, headers=hdr, json=body, timeout=30)
        if r.status_code in (200, 201):
            return True, ""
        return False, f"GitHub書き込みエラー HTTP {r.status_code}: {r.json().get('message', '')}"
    except Exception as e:
        return False, f"通信エラー: {e}"


def save_store(df: pd.DataFrame, action: str) -> tuple[dict, str]:
    """マスタを保存し、更新履歴を記録。GitHub連携があればバックアップも実行。"""
    STORE_DIR.mkdir(exist_ok=True)
    df.to_csv(STORE_CSV, index=False)
    meta: dict = {}
    if META_JSON.exists():
        try:
            meta = json.loads(META_JSON.read_text(encoding="utf-8"))
        except Exception:
            meta = {}
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    meta["last_updated"] = ts
    meta["last_action"] = action
    meta["rows"] = len(df)
    meta.setdefault("history", []).append({"日時": ts, "操作": action, "件数": len(df)})
    meta["history"] = meta["history"][-50:]
    META_JSON.write_text(json.dumps(meta, ensure_ascii=False, indent=1),
                         encoding="utf-8")
    backed, gh_err = _gh_push(action)
    dest = "ローカル＋GitHubバックアップ済み" if backed else "アプリ内ストレージのみ"
    return meta, dest, gh_err

SHEET_NAME = "運用サイトマップ"
HEADER_ROW = 5          # 見出し行（1始まり）
DATA_MIN_ROW = 6        # データ開始行
URL_COL_NAME = "フルURL"
MAX_LEVEL = 6

# TID移行対象_サービスのまとまり整理 v1.0 突合済みのアプリ入口パス
APP_PREFIXES = [
    "/service/", "/member/", "/login", "/profile", "/grade",
    "/cmpn", "/socialfes", "/follow", "/ucar_search", "/webservice",
    "/mailalert_service", "/measurement", "/faq/inquiry",
]
APP_EXACT = {"/recall", "/recall/"}   # /recall はトップのみアプリ（配下は静的届出ページ）

FILE_EXTS = (".pdf", ".jpg", ".jpeg", ".png", ".gif", ".svg", ".zip",
             ".xlsx", ".xls", ".docx", ".doc", ".pptx", ".mp4", ".webp", ".ico")


# ─── 純粋ロジック（テスト可能） ──────────────────────────────

def classify_url(url: str) -> str:
    """URLから HTML（静的） / アプリ（動的） / ファイル を判別する。"""
    try:
        p = urlparse(str(url))
    except Exception:
        return "不明"
    path = p.path or "/"
    low = path.lower()
    if p.netloc and "toyota.jp" not in p.netloc and "toyota-catalog.jp" not in p.netloc:
        return "外部サイト"
    if low.rstrip("/") in APP_EXACT or low in APP_EXACT:
        return "アプリ"
    for pref in APP_PREFIXES:
        if low == pref or low.startswith(pref if pref.endswith("/") else pref + "/") \
           or low.rstrip("/") == pref.rstrip("/"):
            return "アプリ"
    if low.endswith(FILE_EXTS):
        return "ファイル"
    if p.query and not low.endswith((".html", ".htm")):
        return "アプリ"          # クエリ付き動的URL（dc/search 等）
    return "HTML（静的）"


def path_levels(url: str, n: int = MAX_LEVEL) -> list[str]:
    """URLのパスを第1〜第n階層に分解して返す（不足分は空文字）。"""
    try:
        p = urlparse(str(url))
    except Exception:
        return [""] * n
    segs = [s for s in (p.path or "").split("/") if s]
    if segs and segs[-1] in ("index.html", "index.htm"):
        segs = segs[:-1]
    segs = segs[:n]
    return segs + [""] * (n - len(segs))


def norm_url(u: str) -> str:
    """比較用の正規化（フラグメント除去・index.html除去・末尾スラッシュ無視）。"""
    u, _ = urldefrag(str(u).strip())
    u = re.sub(r"/index\.html?$", "/", u)
    u = re.sub(r"^http://", "https://", u)
    return u.rstrip("/").lower()


def load_sitemap(file) -> tuple[pd.DataFrame, list[str]]:
    """運用サイトマップシートを全列DataFrameとして読み込む。"""
    import openpyxl
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"シート「{SHEET_NAME}」が見つかりません")
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=HEADER_ROW, values_only=True))
    raw_headers = list(rows[0])
    headers: list[str] = []
    for i, h in enumerate(raw_headers):
        name = str(h).strip() if h not in (None, "") else f"列{i + 1}"
        while name in headers:
            name += "_"
        headers.append(name)
    data = [list(r) + [None] * (len(headers) - len(r)) for r in rows[1:]]
    df = pd.DataFrame(data, columns=headers)
    if URL_COL_NAME not in df.columns:
        raise ValueError(f"「{URL_COL_NAME}」列が見つかりません")
    # URLが空の行は除外（相対パスは補完）
    df[URL_COL_NAME] = df[URL_COL_NAME].astype("string").str.strip()
    df = df[df[URL_COL_NAME].notna() & (df[URL_COL_NAME] != "")].reset_index(drop=True)
    df[URL_COL_NAME] = df[URL_COL_NAME].map(
        lambda u: "https://toyota.jp" + u if str(u).startswith("/") else u)
    if "転記元" not in "".join(df.columns):
        df["転記元"] = None
    return df, headers


def add_derived(df: pd.DataFrame) -> pd.DataFrame:
    """階層1〜6・階層深さ・種別判定の派生列を付与する。"""
    out = df.copy()
    levels = out[URL_COL_NAME].map(lambda u: path_levels(u))
    for i in range(MAX_LEVEL):
        out[f"階層{i + 1}"] = levels.map(lambda L, i=i: L[i])
    out["階層深さ"] = levels.map(lambda L: sum(1 for s in L if s))
    out["種別判定"] = out[URL_COL_NAME].map(classify_url)
    return out


def agg_directories(dfd: pd.DataFrame, level: int) -> pd.DataFrame:
    """指定階層までのディレクトリ単位に集計する。"""
    keys = [f"階層{i + 1}" for i in range(level)]
    g = dfd.groupby(keys, dropna=False)
    rows = []
    for name, grp in g:
        name = name if isinstance(name, tuple) else (name,)
        path = "/" + "/".join(s for s in name if s)
        if path != "/":
            path += "/"
        rows.append({
            "ディレクトリ": path,
            "ページ数": len(grp),
            "HTML": int((grp["種別判定"] == "HTML（静的）").sum()),
            "アプリ": int((grp["種別判定"] == "アプリ").sum()),
            "その他": int((~grp["種別判定"].isin(["HTML（静的）", "アプリ"])).sum()),
            "代表ページ": str(grp.iloc[0].get("ページ説明") or "")[:40],
        })
    return (pd.DataFrame(rows)
            .sort_values("ページ数", ascending=False)
            .reset_index(drop=True))


def crawl_discover(existing_norm: set[str], start_url: str, path_filter: str,
                   max_pages: int, max_depth: int,
                   progress_cb=None) -> list[dict]:
    """クロールして一覧に無いURLを検知する。toyota.jp内のみ・HTMLのみ巡回。"""
    from fetcher import fetch_html, extract_links, extract_meta
    from config import REQUEST_DELAY

    seen: set[str] = set()
    found: dict[str, dict] = {}
    q: deque[tuple[str, int]] = deque([(start_url, 0)])
    pages = 0
    while q and pages < max_pages:
        url, depth = q.popleft()
        nu = norm_url(url)
        if nu in seen:
            continue
        seen.add(nu)
        try:
            status, html = fetch_html(url)
        except Exception:
            continue
        pages += 1
        if progress_cb:
            progress_cb(pages, max_pages, url, len(found))
        if status != 200 or not html:
            continue
        title = ""
        try:
            title = (extract_meta(html).get("title") or "").strip()
        except Exception:
            pass
        if nu not in existing_norm and nu != norm_url(start_url):
            found[nu] = {"フルURL": url.split("#")[0], "ページ説明": title,
                         "階層深さ": sum(1 for s in path_levels(url) if s),
                         "種別判定": classify_url(url), "リンク元深さ": depth}
        if depth >= max_depth:
            continue
        try:
            links = extract_links(html, url)
        except Exception:
            continue
        for link in links:
            lp = urlparse(link)
            if lp.netloc not in ("toyota.jp", "www.toyota.jp"):
                continue
            if path_filter and not lp.path.startswith(path_filter):
                continue
            if lp.path.lower().endswith(FILE_EXTS):
                continue
            nl = norm_url(link)
            if nl not in seen:
                q.append((link, depth + 1))
        time.sleep(REQUEST_DELAY)
    return sorted(found.values(), key=lambda d: d["フルURL"])


def build_tree(dfd: pd.DataFrame, level: int) -> pd.DataFrame:
    """パス順に並べ、インデント付きツリー列と配下ページ数を持つ表を作る。

    level より深い行は畳む（親の「配下ページ数」に含まれるため情報は失われない）。
    """
    d = dfd.copy()
    d["_segs"] = d[URL_COL_NAME].map(lambda u: tuple(s for s in path_levels(u) if s))
    d = d.sort_values("_segs", kind="stable").reset_index(drop=True)

    # 配下ページ数: 自分のパスを接頭辞とする行数（自分を除く）。
    # 接頭辞カウントで O(n×階層) — 3万行規模でも高速
    from collections import Counter
    seg_list = list(d["_segs"])
    prefix_counts: Counter = Counter()
    exact_counts: Counter = Counter(seg_list)
    for segs in seg_list:
        for i in range(0, len(segs) + 1):
            prefix_counts[segs[:i]] += 1
    d["配下ページ数"] = [prefix_counts[s] - exact_counts[s] for s in seg_list]

    shown = d[d["階層深さ"] <= level].copy()

    def tree_label(row):
        segs = row["_segs"]
        depth = len(segs)
        if depth == 0:
            return "🏠 toyota.jp（トップ）"
        name = segs[-1]
        url = str(row[URL_COL_NAME])
        is_dir = url.rstrip("/").endswith(name) and url.endswith("/")
        icon = "📁 " if (is_dir or row["配下ページ数"] > 0) else "📄 "
        return "　　" * (depth - 1) + "└ " + icon + name + ("/" if is_dir else "")

    shown["ツリー"] = shown.apply(tree_label, axis=1)
    folded = shown["配下ページ数"].where(shown["階層深さ"] == level, 0)
    shown["畳まれた配下"] = folded.map(lambda n: f"+{n}件" if n else "")
    cols = ["ツリー", "配下ページ数", "畳まれた配下", "種別判定", "ページ説明", URL_COL_NAME]
    return shown[[c for c in cols if c in shown.columns]]


ADD_DATE_COL = "追加日"
ADD_SRC_COL = "追加元"
ORPHAN_JSON = STORE_DIR / "orphan_history.json"


def ensure_mgmt_cols(df: pd.DataFrame) -> pd.DataFrame:
    """追加日・追加元の管理列を用意し、既存の転記元表記から日付を復元する。"""
    out = df.copy()
    for c in (ADD_DATE_COL, ADD_SRC_COL):
        if c not in out.columns:
            out[c] = pd.NA
    src_col = next((c for c in out.columns
                    if "転記元" in c and c != ADD_SRC_COL), None)
    if src_col:
        for i, v in out[src_col].items():
            m = re.match(r"クロール検知\s*(\d{4}-\d{2}-\d{2})", str(v or ""))
            if m and pd.isna(out.at[i, ADD_DATE_COL]):
                out.at[i, ADD_DATE_COL] = m.group(1)
                out.at[i, ADD_SRC_COL] = "クロール検知"
    return out


def load_orphan_history() -> list[dict]:
    if ORPHAN_JSON.exists():
        try:
            return json.loads(ORPHAN_JSON.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []


def save_orphan_run(checked: int, errors: int, orphans: list[str]) -> list[dict]:
    hist = load_orphan_history()
    hist.append({"日時": datetime.now().strftime("%Y-%m-%d %H:%M"),
                 "巡回数": checked, "取得失敗": errors, "孤島URL": orphans})
    hist = hist[-20:]
    STORE_DIR.mkdir(exist_ok=True)
    ORPHAN_JSON.write_text(json.dumps(hist, ensure_ascii=False), encoding="utf-8")
    return hist


def diff_orphans(prev: list[str], cur: list[str]) -> tuple[list[str], list[str]]:
    """(新たに孤島になった, 孤島でなくなった)"""
    p, c = set(prev), set(cur)
    return sorted(c - p), sorted(p - c)


def run_orphan_check(master_urls: list[str], fetch_limit: int,
                     progress_cb=None) -> tuple[dict[str, int], int, int]:
    """マスタ全ページを巡回し、マスタ内URLごとの被リンク数を集計する。

    返り値: (正規化URL→被リンク数, 巡回ページ数, 取得失敗数)
    """
    from fetcher import fetch_html, extract_links
    from config import REQUEST_DELAY

    norm_map = {norm_url(u): u for u in master_urls}
    inbound = {n: 0 for n in norm_map}
    sources = [u for u in master_urls
               if not str(u).lower().endswith(FILE_EXTS)][:fetch_limit]
    errors = 0
    for i, u in enumerate(sources, 1):
        if progress_cb:
            progress_cb(i, len(sources), u)
        try:
            status, html = fetch_html(u)
        except Exception:
            errors += 1
            continue
        if status != 200 or not html:
            errors += 1
            continue
        try:
            links = extract_links(html, u)
        except Exception:
            continue
        self_n = norm_url(u)
        for link in set(norm_url(l) for l in links):
            if link in inbound and link != self_n:
                inbound[link] += 1
        time.sleep(REQUEST_DELAY)
    return inbound, len(sources), errors


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=SHEET_NAME)
    return buf.getvalue()


# ─── Streamlit UI ────────────────────────────────────────────

VIEW_COLS = ["№", URL_COL_NAME, "ページ説明", "種別判定", "階層深さ",
             "運用対象コンテンツ", "更新頻度", "情報主幹部署", "運用会社", "アプリ"]


def _cached_derived_impl(_df: pd.DataFrame, ver) -> pd.DataFrame:
    return add_derived(_df)


try:  # マスタ更新時のみ再計算（ver がキー。_df はハッシュ対象外）
    import streamlit as _st
    _cached_derived = _st.cache_data(show_spinner=False)(_cached_derived_impl)
except Exception:  # pragma: no cover
    _cached_derived = _cached_derived_impl


def render():
    import streamlit as st

    st.header("🗂️ サイトマップ管理")
    st.caption("運用サイトマップのマスタをアプリ内に保持し、閲覧・編集・クロール差分検知で更新する管理ツール")

    # 保存後の通知（rerun後も残す）
    if msg := st.session_state.pop("smgr_flash", None):
        st.success(msg)
    if warn := st.session_state.pop("smgr_flash_warn", None):
        st.warning(warn)

    # ── 1. マスタ読み込み（保存済みを自動ロード） ──
    if "smgr_df" not in st.session_state:
        stored, meta = load_store()
        if stored is not None:
            st.session_state["smgr_df"] = stored
            st.session_state["smgr_meta"] = meta

    df: pd.DataFrame | None = st.session_state.get("smgr_df")
    meta: dict = st.session_state.get("smgr_meta", {})

    # ── 取り込み・差し替え（マスタ未登録時のみ展開） ──
    with st.expander("📥 Excelからマスタを取り込み・差し替え", expanded=(df is None)):
        st.caption("tjpコンテンツ管理表.xlsx（運用サイトマップシート）を読み込み、"
                   "アプリ内のマスタとして保存します。以降はアップロード不要です。")
        up = st.file_uploader("tjpコンテンツ管理表.xlsx", type=["xlsx"], key="smgr_upload")
        if up is not None:
            try:
                new_df, _ = load_sitemap(up)
                st.caption(f"読み込みプレビュー: {len(new_df)} 件のURL")
                label = "⚠️ 現在のマスタを差し替えて保存" if df is not None else "💾 マスタとして保存"
                if st.button(label, type="primary", key="smgr_import_btn"):
                    meta, dest, gh_err = save_store(new_df, f"Excel取り込み（{up.name}）")
                    st.session_state["smgr_df"] = new_df
                    st.session_state["smgr_meta"] = meta
                    st.session_state.pop("smgr_new_urls", None)
                    st.session_state["smgr_flash"] = \
                        f"マスタとして保存しました（{len(new_df)}件・{dest}）"
                    if gh_err:
                        st.session_state["smgr_flash_warn"] = \
                            f"GitHubバックアップは失敗しました: {gh_err}"
                    st.rerun()
            except Exception as e:
                st.error(f"読み込みエラー: {e}")

    if df is None:
        st.info("まだマスタが登録されていません。上の「📥 取り込み」から "
                "tjpコンテンツ管理表.xlsx を一度だけ取り込んでください。")
        return

    # ── マスタ状態表示 ──
    gh_on = _gh_cfg() is not None
    st.caption(
        f"🗄️ **マスタ保持中**: {len(df)} 件　｜　最終更新: "
        f"{meta.get('last_updated', 'ー')}（{meta.get('last_action', 'ー')}）　｜　"
        f"保存先: {'アプリ内＋GitHubバックアップ' if gh_on else 'アプリ内ストレージ'}")
    if not gh_on:
        st.warning("GitHubバックアップが未設定です。Streamlit Cloudの再デプロイ時にマスタが"
                   "初期化される可能性があります。Secretsに `GITHUB_TOKEN`（repo権限）を"
                   "設定すると自動バックアップ・自動復元が有効になります。", icon="⚠️")
    hist = meta.get("history", [])
    if hist:
        with st.expander(f"🕘 更新履歴（直近{min(len(hist), 10)}件）"):
            st.dataframe(pd.DataFrame(list(reversed(hist))[:10]),
                         use_container_width=True, hide_index=True)

    # 派生列はマスタ更新時のみ再計算（3万行規模対応）
    _ver = (meta.get("last_updated", ""), len(df))
    dfd = _cached_derived(df, _ver)

    # ── 2. サマリー ──
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("URL数", len(dfd))
    c2.metric("HTML（静的）", int((dfd["種別判定"] == "HTML（静的）").sum()))
    c3.metric("アプリ", int((dfd["種別判定"] == "アプリ").sum()))
    c4.metric("最大階層", int(dfd["階層深さ"].max() or 0))
    c5.metric("第6階層以深", int((dfd["階層深さ"] >= 6).sum()))

    # ── 3. ビュー設定 ──
    st.divider()
    f1, f2, f3, f4 = st.columns([1.4, 1.2, 1.2, 1.4])
    view_mode = f1.radio("表示モード",
                         ["🌲 ツリービュー", "✏️ 一覧編集", "📊 ディレクトリ集計"],
                         key="smgr_mode")
    level = f2.select_slider("表示階層",
                             options=list(range(1, MAX_LEVEL + 1)), value=MAX_LEVEL,
                             format_func=lambda n: f"第{n}階層まで",
                             key="smgr_level")
    dir1 = f3.selectbox(
        "第1階層で絞り込み",
        ["（すべて）"] + sorted(x for x in dfd["階層1"].unique() if x),
        key="smgr_dir1")
    q = f4.text_input("URL・説明で検索", key="smgr_q").strip()

    # 絞り込みは「検索・第1階層」のみ。階層スライダーで行が消えるのはツリーの折り畳みだけ
    view = dfd
    if dir1 != "（すべて）":
        view = view[view["階層1"] == dir1]
    if q:
        mask = (view[URL_COL_NAME].astype(str).str.contains(q, case=False, na=False)
                | view.get("ページ説明", pd.Series("", index=view.index))
                    .astype(str).str.contains(q, case=False, na=False))
        view = view[mask]

    filtered = len(view) != len(dfd)
    fc1, fc2 = st.columns([4, 1])
    fc1.caption(f"全 {len(dfd)} 件中 **{len(view)} 件** を表示"
                + ("（絞り込み中）" if filtered else ""))
    if filtered and fc2.button("絞り込み解除", key="smgr_reset"):
        st.session_state["smgr_dir1"] = "（すべて）"
        st.session_state["smgr_q"] = ""
        st.rerun()

    # ── 4. 表示・編集 ──
    if view_mode == "📊 ディレクトリ集計":
        st.subheader(f"ディレクトリ集計（第{level}階層まで）")
        agg = agg_directories(view, level)
        st.dataframe(agg, use_container_width=True, height=480)
        st.download_button("⬇️ この集計をCSVダウンロード", to_csv_bytes(agg),
                           f"sitemap_dir_level{level}_{date.today()}.csv", "text/csv")
    elif view_mode == "🌲 ツリービュー":
        tree = build_tree(view, level)
        hidden = len(view) - len(tree)
        st.subheader(f"ツリービュー（第{level}階層まで表示・{len(tree)} 行）")
        if hidden:
            st.caption(f"※ 第{level + 1}階層より深い {hidden} 件は畳んでいます"
                       "（「畳まれた配下」列に件数表示。表示階層を上げると展開）")
        st.dataframe(
            tree, use_container_width=True, height=520, hide_index=True,
            column_config={
                "ツリー": st.column_config.TextColumn("サイト構造", width="large"),
                "配下ページ数": st.column_config.NumberColumn("配下", width="small"),
                "畳まれた配下": st.column_config.TextColumn("畳み", width="small"),
                "種別判定": st.column_config.TextColumn("種別", width="small"),
                URL_COL_NAME: st.column_config.LinkColumn("URL", width="medium"),
            },
        )
        st.download_button("⬇️ このツリーをCSV", to_csv_bytes(tree),
                           f"sitemap_tree_level{level}_{date.today()}.csv", "text/csv",
                           key="smgr_tree_csv")
    else:
        pc1, pc2, pc3 = st.columns([1, 1, 2])
        show_all = pc3.checkbox("全列を表示（Excelの全項目）", key="smgr_allcols")
        page_size = pc1.selectbox("1ページの行数", [200, 500, 1000, 2000], index=1,
                                  key="smgr_psize")
        n_pages = max(1, -(-len(view) // page_size))
        page = pc2.number_input(f"ページ（全{n_pages}ページ）", 1, n_pages, 1,
                                key="smgr_page")
        pview = view.iloc[(page - 1) * page_size: page * page_size]
        lvl_cols = [f"階層{i + 1}" for i in range(level)]
        cols = ([c for c in VIEW_COLS if c in view.columns] + lvl_cols) if not show_all \
            else list(view.columns)
        st.subheader(f"一覧編集（{len(view)} 件中 "
                     f"{(page - 1) * page_size + 1}〜{(page - 1) * page_size + len(pview)} 件目）")
        edited = st.data_editor(
            pview[cols], use_container_width=True, height=480,
            num_rows="dynamic", key="smgr_editor",
            disabled=["種別判定", "階層深さ"] + lvl_cols,
        )
        if st.button("💾 編集内容をマスタに保存", type="primary"):
            base = df.copy()
            editable = [c for c in edited.columns
                        if c in base.columns and c not in ("種別判定", "階層深さ")]
            common = edited.index.intersection(pview.index)
            base.loc[common, editable] = edited.loc[common, editable]
            new_rows = edited.loc[edited.index.difference(pview.index), editable]
            new_rows = new_rows[new_rows[URL_COL_NAME].notna()]
            if len(new_rows):
                base = pd.concat([base, new_rows], ignore_index=True)
            meta2, dest, gh_err = save_store(
                base, f"一覧編集（更新{len(common)}行・追加{len(new_rows)}行）")
            st.session_state["smgr_df"] = base
            st.session_state["smgr_meta"] = meta2
            st.session_state["smgr_flash"] = \
                f"マスタに保存しました（更新 {len(common)} 行・追加 {len(new_rows)} 行・{dest}）"
            if gh_err:
                st.session_state["smgr_flash_warn"] = f"GitHubバックアップは失敗しました: {gh_err}"
            st.rerun()

    # ── 5. ダウンロード ──
    d1, d2, d3 = st.columns(3)
    d1.download_button("⬇️ 表示中の絞り込み結果をCSV", to_csv_bytes(view),
                       f"sitemap_view_{date.today()}.csv", "text/csv")
    d2.download_button("⬇️ 全件CSV（判定・階層列つき）", to_csv_bytes(add_derived(df)),
                       f"sitemap_all_{date.today()}.csv", "text/csv")
    d3.download_button("⬇️ Excel（編集反映済み）", to_excel_bytes(df),
                       f"tjpコンテンツ管理表_更新_{date.today()}.xlsx")

    # ── 6. クロール差分検知 ──
    st.divider()
    with st.expander("🕷️ クロールして一覧に無いURLを検知・追加", expanded=False):
        cc1, cc2, cc3, cc4 = st.columns([1.6, 1.2, 1, 1])
        start_url = cc1.text_input("開始URL", "https://toyota.jp/", key="smgr_c_start")
        path_filter = cc2.text_input("パス制限（任意）", "", placeholder="/carlineup/",
                                     key="smgr_c_filter").strip()
        max_pages = cc3.number_input("最大巡回ページ数", 10, 2000, 200, step=10,
                                     key="smgr_c_max")
        max_depth = cc4.number_input("リンク深さ", 1, MAX_LEVEL, 3, key="smgr_c_depth")
        st.caption("toyota.jp内のHTMLのみ巡回します。一覧との差分（未登録URL）を検知します。")

        if st.button("▶ クロール開始", key="smgr_c_run"):
            existing = set(dfd[URL_COL_NAME].map(norm_url))
            bar = st.progress(0.0)
            info = st.empty()

            def cb(done, total, url, nfound):
                bar.progress(min(done / total, 1.0))
                info.caption(f"{done}/{total} ページ巡回中… 未登録検知 {nfound} 件　{url[:80]}")

            new_urls = crawl_discover(existing, start_url, path_filter,
                                      int(max_pages), int(max_depth), cb)
            st.session_state["smgr_new_urls"] = new_urls
            bar.progress(1.0)

        new_urls = st.session_state.get("smgr_new_urls")
        if new_urls is not None:
            if not new_urls:
                st.success("一覧に無いURLは見つかりませんでした。")
            else:
                st.warning(f"一覧に無いURLを {len(new_urls)} 件検知しました。"
                           "追加する行にチェックを入れてください。")
                nd = pd.DataFrame(new_urls)
                nd.insert(0, "追加", True)
                sel = st.data_editor(nd, use_container_width=True, height=320,
                                     key="smgr_new_editor",
                                     disabled=[c for c in nd.columns if c != "追加"])
                a1, a2 = st.columns([1, 2])
                if a1.button("➕ チェックした行を一覧に追加", type="primary",
                             key="smgr_add_btn"):
                    add = sel[sel["追加"]]
                    base = ensure_mgmt_cols(df)
                    src_col = next((c for c in base.columns
                                    if "転記元" in c and c != ADD_SRC_COL), None)
                    rows = []
                    for _, r in add.iterrows():
                        row = {c: None for c in base.columns}
                        row[URL_COL_NAME] = r["フルURL"]
                        if "ページ説明" in base.columns:
                            row["ページ説明"] = r["ページ説明"]
                        if src_col:
                            row[src_col] = f"クロール検知 {date.today()}"
                        row[ADD_DATE_COL] = str(date.today())
                        row[ADD_SRC_COL] = "クロール検知"
                        rows.append(row)
                    base = pd.concat([base, pd.DataFrame(rows)], ignore_index=True)
                    meta2, dest, gh_err = save_store(base, f"クロール検知 {len(rows)}件追加")
                    st.session_state["smgr_df"] = base
                    st.session_state["smgr_meta"] = meta2
                    st.session_state["smgr_new_urls"] = [
                        u for u in new_urls
                        if u["フルURL"] not in set(add["フルURL"])]
                    st.session_state["smgr_flash"] = \
                        f"{len(rows)} 件をマスタに追加・保存しました（{dest}）"
                    if gh_err:
                        st.session_state["smgr_flash_warn"] = \
                            f"GitHubバックアップは失敗しました: {gh_err}"
                    st.rerun()
                a2.download_button("⬇️ 検知結果をCSV", to_csv_bytes(nd.drop(columns=['追加'])),
                                   f"crawl_new_urls_{date.today()}.csv", "text/csv",
                                   key="smgr_new_csv")

    # ── 7. 追加URLの差分管理 ──
    with st.expander("📅 追加URLの差分管理（クロール検知で増えた行）", expanded=False):
        dm = ensure_mgmt_cols(df)
        added = dm[dm[ADD_SRC_COL].notna()]
        if added.empty:
            st.info("クロール検知で追加されたURLはまだありません。")
        else:
            bydate = (added.groupby(ADD_DATE_COL).size()
                      .rename("追加件数").reset_index()
                      .sort_values(ADD_DATE_COL, ascending=False))
            st.dataframe(bydate, hide_index=True, use_container_width=True)
            days = st.multiselect("追加日で絞り込み",
                                  sorted(added[ADD_DATE_COL].dropna().unique(),
                                         reverse=True),
                                  key="smgr_diff_days")
            shown = added if not days else added[added[ADD_DATE_COL].isin(days)]
            cols = [c for c in ("№", URL_COL_NAME, "ページ説明",
                                ADD_DATE_COL, ADD_SRC_COL) if c in shown.columns]
            view2 = shown[cols].copy()
            view2.insert(0, "取り消し", False)
            sel2 = st.data_editor(view2, use_container_width=True, height=300,
                                  key="smgr_diff_editor",
                                  disabled=[c for c in view2.columns if c != "取り消し"])
            b1, b2 = st.columns([1, 2])
            if b1.button("🗑️ チェックした行をマスタから削除", key="smgr_diff_del"):
                drop_idx = sel2[sel2["取り消し"]].index
                if len(drop_idx):
                    base = ensure_mgmt_cols(df).drop(index=drop_idx).reset_index(drop=True)
                    meta2, dest, gh_err = save_store(
                        base, f"差分管理で削除 {len(drop_idx)}件")
                    st.session_state["smgr_df"] = base
                    st.session_state["smgr_meta"] = meta2
                    st.session_state["smgr_flash"] = \
                        f"{len(drop_idx)} 件をマスタから削除しました（{dest}）"
                    if gh_err:
                        st.session_state["smgr_flash_warn"] = \
                            f"GitHubバックアップは失敗しました: {gh_err}"
                    st.rerun()
            b2.download_button("⬇️ 追加分をCSV", to_csv_bytes(shown[cols]),
                               f"added_urls_{date.today()}.csv", "text/csv",
                               key="smgr_diff_csv")

    # ── 8. 陸の孤島チェック ──
    with st.expander("🏝️ 陸の孤島チェック（どこからもリンクされていないページ）",
                     expanded=False):
        st.caption("マスタの各ページを巡回してリンクを収集し、マスタ内のどのページからも"
                   "リンクされていないURL（陸の孤島）を検出します。"
                   "※ JavaScriptで動的に生成されるリンクは検出できないため、"
                   "孤島判定は【要検証】として扱ってください。")
        o1, o2 = st.columns([1, 2])
        fetch_limit = o1.number_input("巡回ページ数の上限", 50, 30000,
                                      min(len(df), 1000), step=50,
                                      key="smgr_o_limit")
        o2.caption(f"目安時間: 約{int(fetch_limit) * 0.5 / 60:.0f}分"
                   "（1ページ約0.5秒）。上限をマスタ件数以上にすると全ページ巡回。")
        if st.button("▶ 孤島チェック実行", key="smgr_o_run"):
            urls = [str(u) for u in df[URL_COL_NAME].dropna()
                    if str(u).startswith("http")]
            bar = st.progress(0.0)
            info = st.empty()
            inbound, checked, errors = run_orphan_check(
                urls, int(fetch_limit),
                lambda i, n, u: (bar.progress(i / n),
                                 info.caption(f"{i}/{n} 巡回中… {u[:80]}")))
            nm = {norm_url(u): u for u in urls}
            orphans = [nm[k] for k, v in inbound.items() if v == 0]
            hist = save_orphan_run(checked, errors, orphans)
            st.session_state["smgr_orphan"] = {
                "inbound": {nm[k]: v for k, v in inbound.items()},
                "orphans": orphans, "checked": checked, "errors": errors,
                "hist": hist}
            bar.progress(1.0)

        res = st.session_state.get("smgr_orphan")
        if res is None:
            hist = load_orphan_history()
            if hist:
                last = hist[-1]
                st.caption(f"前回実行: {last['日時']}（巡回{last['巡回数']}件・"
                           f"孤島{len(last['孤島URL'])}件）")
        else:
            m1, m2, m3 = st.columns(3)
            m1.metric("巡回ページ数", res["checked"])
            m2.metric("🏝️ 孤島URL", len(res["orphans"]))
            m3.metric("取得失敗", res["errors"])
            hist = res["hist"]
            if len(hist) >= 2:
                new_o, fixed_o = diff_orphans(hist[-2]["孤島URL"], res["orphans"])
                st.caption(f"前回（{hist[-2]['日時']}）との差分: "
                           f"新たに孤島 {len(new_o)}件 ／ 解消 {len(fixed_o)}件")
                if new_o:
                    st.error("🆕 新たに孤島になったURL:\n" +
                             "\n".join(f"- {u}" for u in new_o[:20]))
                if fixed_o:
                    st.success("✅ 孤島が解消されたURL:\n" +
                               "\n".join(f"- {u}" for u in fixed_o[:20]))
            od = dfd[[URL_COL_NAME, "ページ説明", "種別判定"]].copy()
            od["被リンク数"] = od[URL_COL_NAME].map(
                lambda u: res["inbound"].get(str(u), pd.NA))
            od["判定"] = od["被リンク数"].map(
                lambda v: "🏝️ 孤島" if v == 0 else ("" if pd.isna(v) else "リンクあり"))
            only_o = st.checkbox("孤島のみ表示", value=True, key="smgr_o_only")
            show = od[od["判定"] == "🏝️ 孤島"] if only_o else od
            st.dataframe(show.sort_values("被リンク数", na_position="last"),
                         use_container_width=True, height=360, hide_index=True)
            st.download_button("⬇️ 孤島チェック結果をCSV", to_csv_bytes(od),
                               f"orphan_check_{date.today()}.csv", "text/csv",
                               key="smgr_o_csv")
