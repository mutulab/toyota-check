"""表記ゆれ辞書ローダー（Excelベース）"""
from __future__ import annotations
import re
from pathlib import Path

_TMP_PATH = Path("/tmp/toyota-check-hyoki.xlsx")
_BUNDLED_PATH = Path(__file__).parent / "hyoki_dict.xlsx"

_FALLBACK: dict[str, str] = {
    r"Webサイト|web\s*サイト|ウェブサイト": "WEBサイト",
    r"E-?mail|e-?mail|メール": "メール",
    r"お問い合わせ|お問合せ|お問合わせ": "お問い合わせ",
    r"ログイン|ろぐいん|log\s*in": "ログイン",
    r"ホームページ": "ホームページ（※WEBサイトに統一検討）",
}


def _find_sheet(wb):
    for name in wb.sheetnames:
        if "用語リスト" in name or "用語" in name:
            return wb[name]
    return wb.active


def read_excel_flat(source) -> dict[str, str]:
    """Excel → {NG表記: 推奨表記} の平文辞書
    Row 1-2: 空白, Row 3: ヘッダー（スキップ）, Row 4+: データ
    G列(index 6)=推奨表記（複数行は1行目を使用）
    H列(index 7)=NG表記（「、」か改行区切り。「-」は対象外）
    """
    import openpyxl
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws = _find_sheet(wb)
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 8:
            continue
        correct_raw, wrong_raw = row[6], row[7]
        if not correct_raw or not wrong_raw:
            continue
        wrong_s = str(wrong_raw).strip()
        if wrong_s in ("-", "—", "－", ""):
            continue  # section header rows have "-" as NG form
        # Take first line of correct form if multiline
        c = str(correct_raw).strip().splitlines()[0].strip()
        if not c:
            continue
        # Split NG forms by 「、」 and newlines
        for sep_part in wrong_s.replace("、", "\n").splitlines():
            w = sep_part.strip()
            if w and w != c and w not in ("-", "—", "－"):
                result[w] = c
    return result


def load_flat() -> dict[str, str]:
    """表示・差分用の平文辞書（NG表記 → 推奨表記）"""
    for path in (_TMP_PATH, _BUNDLED_PATH):
        if path.exists():
            try:
                return read_excel_flat(path)
            except Exception:
                pass
    return {}


def load_for_check() -> dict[str, str]:
    """チェック用: {regex_pattern: 推奨表記}。Excel辞書優先、なければハードコード"""
    flat = load_flat()
    if flat:
        return {re.escape(k): v for k, v in flat.items()}
    return dict(_FALLBACK)


def save_override(data: bytes) -> None:
    """アップロードされたExcelを /tmp に保存して次回から使用"""
    _TMP_PATH.write_bytes(data)


def source_label() -> str:
    if _TMP_PATH.exists():
        return "カスタム辞書（更新済）"
    if _BUNDLED_PATH.exists():
        return "デフォルト辞書（toyota_lexus用語リスト）"
    return "ハードコード辞書（フォールバック）"
